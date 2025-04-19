# Full working script incorporating Summary Reports, Basic Dashboard,
# plus previous features (Run ID, CSV logging, log viewer, themes, window persistence,
# worklist, delta checks, batch import, history viewer, flowchart viewer, about box,
# system info, self-test, config validation)
# Version 1.10 - Complete

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import sys
import os
from datetime import datetime, timedelta # Added timedelta for dashboard default range
import logging
import sqlite3 # Built-in database library
import csv     # Built-in CSV handling
import json    # For configuration
import time    # For barcode timing simulation
import platform # For system info
import traceback # For detailed error logging
from collections import Counter # For counting indeterminate reasons

# --- Required External Libraries ---
# Try importing optional libraries for export/display features
try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT # For alignment constants
    REPORTLAB_AVAILABLE = True
except ImportError: REPORTLAB_AVAILABLE = False; print("Warning: reportlab not found. PDF export disabled.", file=sys.stderr)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError: OPENPYXL_AVAILABLE = False; print("Warning: openpyxl not found. Excel import/export disabled.", file=sys.stderr)

try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError: PIL_AVAILABLE = False; print("Warning: Pillow (PIL) not found. Flowchart display disabled.", file=sys.stderr)


# --- Constants ---
APP_VERSION = "1.10" # Incremented version
APP_TITLE = f"LIAISON® QFT®-Plus Interpreter v{APP_VERSION}"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FLOWCHART_PATH = os.path.join(SCRIPT_DIR, "qft_flowchart.png")
LOG_FILENAME = os.path.join(SCRIPT_DIR, "qft_interpreter_log.csv")
DB_FILENAME = os.path.join(SCRIPT_DIR, "qft_history.db")
CONFIG_FILENAME = os.path.join(SCRIPT_DIR, "qft_config.json")

REQUIRED_BATCH_HEADERS = {'sample id', 'nil', 'tb1', 'tb2', 'mitogen'}
LOG_HEADER = ["Timestamp", "OperatorID", "RunID", "SampleID", "Nil", "TB1", "TB2", "Mitogen", "Result", "Reason"]
DEFAULT_CONFIG = {"geometry": "700x750", "theme": "clam", "dashboard_days": 7}
CONFIG_KEYS = {"geometry", "theme", "dashboard_days"}

BARCODE_INTERKEY_DELAY = 0.1; BARCODE_MIN_LEN = 3
SELF_TEST_CASES = [("Clear Positive (TB1)", 0.10, 1.50, 0.20, 5.0, "POS†"),("Clear Positive (TB2)", 0.20, 0.40, 2.00, 6.0, "POS†"),("Clear Negative", 0.10, 0.20, 0.30, 2.0, "NEG"),("Indeterminate (High Nil)", 9.50, 10.0, 11.0, 15.0, "IND*"),("Indeterminate (Low Mitogen)", 0.20, 0.30, 0.40, 0.60, "IND*"),("Borderline Positive (TB1 near 0.35)", 0.10, 0.45, 0.20, 3.0, "POS†"),("Borderline Negative (TB1 below 0.35)", 0.10, 0.40, 0.20, 3.0, "NEG"),("Borderline Positive (TB1 meets 25% rule)", 1.00, 1.35, 0.50, 4.0, "POS†"), ("Borderline Negative (TB1 fails 25% rule)", 1.60, 1.95, 0.50, 4.0, "NEG")]

# --- Logging Setup ---
def setup_logging():
    """Sets up CSV logging, adding header if file is new/empty."""
    log_exists = os.path.exists(LOG_FILENAME); log_empty = (log_exists and os.path.getsize(LOG_FILENAME) == 0)
    log_formatter = logging.Formatter('%(message)s'); log_handler = logging.FileHandler(LOG_FILENAME, mode='a', encoding='utf-8'); log_handler.setFormatter(log_formatter)
    logger = logging.getLogger('QFTLogger'); logger.setLevel(logging.INFO)
    if not logger.handlers: logger.addHandler(log_handler)
    if not log_exists or log_empty:
        try:
            with open(LOG_FILENAME, 'a', newline='', encoding='utf-8') as f: csv.writer(f).writerow(LOG_HEADER)
        except IOError as e: print(f"Error writing log header: {e}", file=sys.stderr)
    return logger
qft_logger = setup_logging()

def log_event(event_type="INFO", op_id="System", run_id="N/A", sample_id="N/A", result="N/A", reason="N/A", details=""):
    """Writes a structured event to the CSV log file."""
    if event_type != "INTERPRET": return # Only log interpretations for now
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        nil_val = details.get('nil',''); tb1_val = details.get('tb1',''); tb2_val = details.get('tb2',''); mit_val = details.get('mit','')
        try: nil_str = f"{float(nil_val):.3f}" if nil_val else ""
        except: nil_str = str(nil_val)
        try: tb1_str = f"{float(tb1_val):.3f}" if tb1_val else ""
        except: tb1_str = str(tb1_val)
        try: tb2_str = f"{float(tb2_val):.3f}" if tb2_val else ""
        except: tb2_str = str(tb2_val)
        try: mit_str = f"{float(mit_val):.3f}" if mit_val else ""
        except: mit_str = str(mit_val)
        log_row = [timestamp, op_id, run_id, sample_id, nil_str, tb1_str, tb2_str, mit_str, result, reason]
        with open(LOG_FILENAME, 'a', newline='', encoding='utf-8') as f: csv.writer(f).writerow(log_row)
    except Exception as e: print(f"Failed to write to log file: {e}", file=sys.stderr)


# --- Configuration Handling ---
def load_config():
    """Loads configuration from JSON file with validation."""
    config = DEFAULT_CONFIG.copy();
    if not os.path.exists(CONFIG_FILENAME): print(f"Info: Config file '{CONFIG_FILENAME}' not found. Using defaults."); return config
    try:
        with open(CONFIG_FILENAME, 'r') as f: loaded_config = json.load(f)
        valid_config = {k: v for k, v in loaded_config.items() if k in CONFIG_KEYS}
        missing_keys = CONFIG_KEYS - set(valid_config.keys())
        if missing_keys: print(f"Warning: Config missing keys: {missing_keys}. Using defaults.")
        config.update(valid_config); print("Info: Configuration loaded."); return config
    except (json.JSONDecodeError, IOError) as e: print(f"Warning: Could not load config '{CONFIG_FILENAME}': {e}. Using defaults.", file=sys.stderr); return DEFAULT_CONFIG.copy()

def save_config(config_data):
    """Saves configuration to JSON file."""
    try:
        with open(CONFIG_FILENAME, 'w') as f: json.dump(config_data, f, indent=4); print("Info: Configuration saved.")
    except IOError as e: print(f"Warning: Could not save config '{CONFIG_FILENAME}': {e}", file=sys.stderr)


# --- Database Setup and Helpers ---
def init_db():
    """Initializes the SQLite database and adds run_id column if needed."""
    try:
        conn = sqlite3.connect(DB_FILENAME); cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS interpretations (id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT NOT NULL, operator_id TEXT, sample_id TEXT NOT NULL, nil_value REAL, tb1_value REAL, tb2_value REAL, mit_value REAL, result TEXT, reason TEXT, run_id TEXT)''')
        cursor.execute("PRAGMA table_info(interpretations)"); columns = [info[1] for info in cursor.fetchall()]
        if 'run_id' not in columns: print("Upgrading DB: Adding 'run_id' column."); cursor.execute('ALTER TABLE interpretations ADD COLUMN run_id TEXT')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_sample_id ON interpretations (sample_id)'); cursor.execute('CREATE INDEX IF NOT EXISTS idx_timestamp ON interpretations (timestamp)'); cursor.execute('CREATE INDEX IF NOT EXISTS idx_run_id ON interpretations (run_id)')
        conn.commit(); conn.close(); print(f"Database '{DB_FILENAME}' initialized/verified.")
    except sqlite3.Error as e: print(f"DB Init/Upgrade Error: {e}", file=sys.stderr); messagebox.showerror("Database Error", f"Could not initialize/upgrade history database:\n{e}")

def save_interpretation_to_db(data):
    """Saves a result dictionary (including run_id) to the database."""
    conn = None
    try:
        conn = sqlite3.connect(DB_FILENAME); cursor = conn.cursor()
        cursor.execute('''INSERT INTO interpretations (timestamp, operator_id, sample_id, run_id, nil_value, tb1_value, tb2_value, mit_value, result, reason) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                       (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), data.get('operator_id', 'N/A'), data.get('sample_id', 'N/A'), data.get('run_id', 'N/A'), data.get('input_nil'), data.get('input_tb1'), data.get('input_tb2'), data.get('input_mit'), data.get('result'), data.get('reason')))
        conn.commit(); return True
    except sqlite3.Error as e: print(f"DB Save Error: {e}", file=sys.stderr); messagebox.showerror("Database Error", f"Failed to save result to history for {data.get('sample_id','N/A')}:\n{e}"); return False
    except Exception as e: print(f"Unexpected error during DB save: {e}", file=sys.stderr); traceback.print_exc(); messagebox.showerror("Save Error", f"Unexpected error saving history:\n{e}"); return False
    finally:
        if conn: conn.close()

def query_db_for_reports(start_date_str, end_date_str):
    """Queries DB for results within a date range for reporting."""
    conn = None; results = []
    try:
        start_dt = datetime.strptime(start_date_str, '%Y-%m-%d').strftime('%Y-%m-%d 00:00:00')
        end_dt = datetime.strptime(end_date_str, '%Y-%m-%d').strftime('%Y-%m-%d 23:59:59')
        conn = sqlite3.connect(DB_FILENAME); cursor = conn.cursor()
        cursor.execute("SELECT timestamp, operator_id, run_id, sample_id, result, reason FROM interpretations WHERE timestamp BETWEEN ? AND ? ORDER BY timestamp ASC", (start_dt, end_dt))
        results = cursor.fetchall(); return results
    except sqlite3.Error as e: messagebox.showerror("Database Error", f"Failed to query history for report:\n{e}"); return []
    except ValueError: messagebox.showerror("Date Error", "Invalid date format. Please use YYYY-MM-DD."); return []
    finally:
        if conn: conn.close()

# --- Core Interpretation Logic ---
def interpret_qft(nil, tb1, tb2, mit):
    """ Interprets QFT results, returns a dictionary. """
    tb1_minus_nil=tb1-nil; tb2_minus_nil=tb2-nil; mit_minus_nil=mit-nil; nil_25_percent=0.25*nil if nil>=0 else 0; qft_result=""; reason=""
    if nil>8.0: qft_result="IND*"; reason=f"High Nil Control ({nil:.3f} > 8.0 IU/mL)"
    else:
        is_tb1_pos=tb1_minus_nil>=0.35 and tb1_minus_nil>=nil_25_percent; is_tb2_pos=tb2_minus_nil>=0.35 and tb2_minus_nil>=nil_25_percent
        if is_tb1_pos: qft_result="POS†"; reason=f"TB1 Antigen positive (TB1-Nil={tb1_minus_nil:.3f} IU/mL)"
        elif is_tb2_pos: qft_result="POS†"; reason=f"TB2 Antigen positive (TB2-Nil={tb2_minus_nil:.3f} IU/mL)"
        else:
            if mit_minus_nil>=0.5: qft_result="NEG"; reason="TB Antigens negative, Mitogen control valid"
            else: qft_result="IND*"; reason=f"Low Mitogen Control (Mit-Nil={mit_minus_nil:.3f} < 0.5 IU/mL difference)"
    return {"result":qft_result, "tb1_nil":tb1_minus_nil, "tb2_nil":tb2_minus_nil, "mit_nil":mit_minus_nil, "nil_25":nil_25_percent, "reason":reason, "input_nil":nil, "input_tb1":tb1, "input_tb2":tb2, "input_mit":mit}

def get_previous_result(sample_id):
    """Queries DB for the most recent result for a given Sample ID."""
    conn=None;
    try: conn=sqlite3.connect(DB_FILENAME); cursor=conn.cursor(); cursor.execute("SELECT result, timestamp FROM interpretations WHERE sample_id = ? ORDER BY timestamp DESC LIMIT 1", (sample_id,)); result=cursor.fetchone(); return result
    except sqlite3.Error as e: log_event("ERROR", sample_id=sample_id, details=f"Delta Check DB Error: {e}"); return None
    finally:
        if conn: conn.close()

def check_significant_change(prev_res, curr_res):
    """Determines if a result change is significant for Delta Check."""
    if prev_res == curr_res: return False
    significant_pairs={("NEG", "POS†"),("POS†", "NEG"),("NEG", "IND*"),("POS†", "IND*")};
    if (prev_res, curr_res) in significant_pairs or (curr_res, prev_res) in significant_pairs: return True
    if prev_res == "IND*" and curr_res in ("POS†", "NEG"): return True
    if curr_res == "IND*" and prev_res in ("POS†", "NEG"): return True
    return False

# --- GUI Application Class ---
class QFTApp:
    def __init__(self, master):
        self.master = master
        self.config = load_config()
        self.is_processing = False
        init_db()
        log_event("INFO", details=f"Application Started. Version: {APP_VERSION}")

        master.title(APP_TITLE)
        master.geometry(self.config.get("geometry", "700x750")) # Use loaded geometry

        # Styles and Theme Setup
        self.style = ttk.Style(); self.available_themes = self.style.theme_names()
        self.current_theme = tk.StringVar(value=self.config.get("theme", "clam"))
        try:
            if self.current_theme.get() not in self.available_themes: self.current_theme.set("clam")
            self.style.theme_use(self.current_theme.get())
        except: self.current_theme.set(self.style.theme_use());
        self._configure_styles()

        self.last_results = None; self.clipboard_content = tk.StringVar()
        self.worklist_items = []
        self._barcode_buffer = ""; self._last_key_time = 0

        # --- Menu Bar ---
        self.menu_bar = tk.Menu(master); master.config(menu=self.menu_bar)
        self.view_menu = tk.Menu(self.menu_bar, tearoff=0); self.menu_bar.add_cascade(label="View", menu=self.view_menu)
        self.theme_menu = tk.Menu(self.view_menu, tearoff=0); self.view_menu.add_cascade(label="Theme", menu=self.theme_menu)
        for theme in self.available_themes: self.theme_menu.add_radiobutton(label=theme, variable=self.current_theme, value=theme, command=self.change_theme)
        self.tools_menu = tk.Menu(self.menu_bar, tearoff=0); self.menu_bar.add_cascade(label="Tools", menu=self.tools_menu)
        self.tools_menu.add_command(label="Worklist Manager...", command=self.show_worklist_window); self.tools_menu.add_command(label="View Log File...", command=self.show_log_viewer_window); self.tools_menu.add_command(label="Run Self-Test...", command=self.run_self_test); self.tools_menu.add_command(label="Generate Summary Report...", command=self.show_report_window); self.tools_menu.add_command(label="Refresh Dashboard", command=self.update_dashboard)
        self.history_menu = tk.Menu(self.menu_bar, tearoff=0); self.menu_bar.add_cascade(label="History", menu=self.history_menu)
        self.history_menu.add_command(label="View Interpretation History...", command=self.show_history_window)
        self.help_menu = tk.Menu(self.menu_bar, tearoff=0); self.menu_bar.add_cascade(label="Help", menu=self.help_menu)
        self.help_menu.add_command(label="Batch Import Format...", command=self.show_batch_format_help); self.help_menu.add_command(label="View Flowchart", command=self.show_flowchart)
        if not (PIL_AVAILABLE and os.path.exists(FLOWCHART_PATH)): self.help_menu.entryconfig("View Flowchart", state=tk.DISABLED)
        self.help_menu.add_separator(); self.help_menu.add_command(label="About / System Info...", command=self.show_about)

        # --- Dashboard Frame (NEW) ---
        self.dashboard_frame = ttk.LabelFrame(master, text="Dashboard (Last {} Days)".format(self.config.get('dashboard_days', 7)), padding="10")
        self.dashboard_frame.grid(row=0, column=0, padx=10, pady=(10,0), sticky="ew")
        self.dashboard_vars = {"total": tk.StringVar(value="Total: -"), "pos": tk.StringVar(value="POS: -"), "neg": tk.StringVar(value="NEG: -"), "ind": tk.StringVar(value="IND: -"), "ind_high_nil": tk.StringVar(value="IND (High Nil): -"), "ind_low_mit": tk.StringVar(value="IND (Low Mit): -"), "pos_rate": tk.StringVar(value="POS%: -"), "ind_rate": tk.StringVar(value="IND%: -")}
        ttk.Label(self.dashboard_frame, textvariable=self.dashboard_vars["total"]).grid(row=0, column=0, sticky="w", padx=5); ttk.Label(self.dashboard_frame, textvariable=self.dashboard_vars["pos"]).grid(row=0, column=1, sticky="w", padx=5); ttk.Label(self.dashboard_frame, textvariable=self.dashboard_vars["neg"]).grid(row=0, column=2, sticky="w", padx=5); ttk.Label(self.dashboard_frame, textvariable=self.dashboard_vars["ind"]).grid(row=0, column=3, sticky="w", padx=5)
        ttk.Label(self.dashboard_frame, textvariable=self.dashboard_vars["ind_high_nil"]).grid(row=1, column=0, columnspan=2, sticky="w", padx=5); ttk.Label(self.dashboard_frame, textvariable=self.dashboard_vars["ind_low_mit"]).grid(row=1, column=2, columnspan=2, sticky="w", padx=5)
        ttk.Label(self.dashboard_frame, textvariable=self.dashboard_vars["pos_rate"]).grid(row=0, column=4, sticky="w", padx=(15, 5)); ttk.Label(self.dashboard_frame, textvariable=self.dashboard_vars["ind_rate"]).grid(row=1, column=4, sticky="w", padx=(15, 5))
        self.master.after(500, self.update_dashboard) # Initial dashboard update


        # --- Input Frame ---
        input_frame = ttk.Frame(master, padding="15 15 15 10"); input_frame.grid(row=1, column=0, sticky="nsew") # Shifted down
        ttk.Label(input_frame, text="Run/Sample Information & Raw Values", style='Header.TLabel').grid(row=0, column=0, columnspan=4, pady=(0, 10), sticky='w')
        ttk.Label(input_frame, text="Operator ID:").grid(row=1, column=0, sticky="w", padx=5, pady=5); self.op_id_entry = ttk.Entry(input_frame, width=10); self.op_id_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        ttk.Label(input_frame, text="Run ID:").grid(row=1, column=2, sticky="w", padx=(10, 5), pady=5); self.run_id_entry = ttk.Entry(input_frame, width=15); self.run_id_entry.grid(row=1, column=3, padx=5, pady=5, sticky="w")
        ttk.Label(input_frame, text="Sample ID:").grid(row=2, column=0, sticky="w", padx=5, pady=5); self.sample_id_entry = ttk.Entry(input_frame, width=25); self.sample_id_entry.grid(row=2, column=1, columnspan=3, padx=5, pady=5, sticky="ew")
        ttk.Label(input_frame, text="Nil Control:").grid(row=3, column=0, sticky="w", padx=5, pady=5); self.nil_entry = ttk.Entry(input_frame, width=15, justify='right'); self.nil_entry.grid(row=3, column=1, padx=5, pady=5); ttk.Label(input_frame, text="IU/mL").grid(row=3, column=2, sticky="w", padx=2)
        ttk.Label(input_frame, text="TB Antigen 1 (TB1):").grid(row=4, column=0, sticky="w", padx=5, pady=5); self.tb1_entry = ttk.Entry(input_frame, width=15, justify='right'); self.tb1_entry.grid(row=4, column=1, padx=5, pady=5); ttk.Label(input_frame, text="IU/mL").grid(row=4, column=2, sticky="w", padx=2)
        ttk.Label(input_frame, text="TB Antigen 2 (TB2):").grid(row=5, column=0, sticky="w", padx=5, pady=5); self.tb2_entry = ttk.Entry(input_frame, width=15, justify='right'); self.tb2_entry.grid(row=5, column=1, padx=5, pady=5); ttk.Label(input_frame, text="IU/mL").grid(row=5, column=2, sticky="w", padx=2)
        ttk.Label(input_frame, text="Mitogen Control (Mit):").grid(row=6, column=0, sticky="w", padx=5, pady=5); self.mit_entry = ttk.Entry(input_frame, width=15, justify='right'); self.mit_entry.grid(row=6, column=1, padx=5, pady=5); ttk.Label(input_frame, text="IU/mL").grid(row=6, column=2, sticky="w", padx=2)
        self.op_id_entry.focus_set()
        self.mit_entry.bind("<Return>", self.run_interpretation)
        master.bind('<Key>', self.handle_key_event)

        # --- Button Frame ---
        button_frame = ttk.Frame(master, padding="15 5 15 10"); button_frame.grid(row=2, column=0, sticky="ew") # Shifted down
        self.interpret_button = ttk.Button(button_frame, text="Interpret Single", command=self.run_interpretation); self.interpret_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.import_button = ttk.Button(button_frame, text="Import Batch...", command=self.import_batch); self.import_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.clear_button = ttk.Button(button_frame, text="Clear", command=self.clear_fields); self.clear_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.copy_button = ttk.Button(button_frame, text="Copy Summary", command=self.copy_summary_to_clipboard, state=tk.DISABLED); self.copy_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.export_pdf_button = ttk.Button(button_frame, text="Export PDF", command=self.export_to_pdf, state=tk.DISABLED)
        if REPORTLAB_AVAILABLE: self.export_pdf_button.pack(side=tk.RIGHT, padx=5, pady=5)
        self.export_excel_button = ttk.Button(button_frame, text="Export Excel", command=self.export_to_excel, state=tk.DISABLED)
        if OPENPYXL_AVAILABLE: self.export_excel_button.pack(side=tk.RIGHT, padx=5, pady=5)

        # --- Results Frame ---
        results_outer_frame = ttk.Frame(master, padding="15 5 15 15"); results_outer_frame.grid(row=3, column=0, sticky="nsew") # Shifted down
        self.result_label_var = tk.StringVar(value="---"); self.result_label = ttk.Label(results_outer_frame, textvariable=self.result_label_var, style='ResultDefault.TLabel', anchor='center'); self.result_label.pack(fill=tk.X, pady=(5, 10))
        ttk.Label(results_outer_frame, text="Interpretation Detail:", style='Header.TLabel').pack(anchor='w', pady=(0, 5))
        self.results_text = tk.Text(results_outer_frame, height=15, width=70, wrap=tk.WORD, font=('Consolas', 9), state='disabled', borderwidth=1, relief="sunken"); self.results_text.pack(fill=tk.BOTH, expand=True)
        self.results_text.tag_configure("pos_ind_result", foreground="red", font=('Consolas', 9, 'bold')); self.results_text.tag_configure("neg_result", foreground="darkgreen", font=('Consolas', 9, 'bold'))

        # --- Status Bar ---
        self.status_var = tk.StringVar(); self.status_bar = ttk.Label(master, textvariable=self.status_var, style="Status.TLabel", relief=tk.SUNKEN, anchor=tk.W); self.status_bar.grid(row=4, column=0, sticky='ew') # Shifted down
        self.set_status("Ready")

        # Configure row/column weights
        master.grid_rowconfigure(3, weight=1) # Results area expands (index changed)
        master.grid_columnconfigure(0, weight=1)
        results_outer_frame.grid_rowconfigure(2, weight=1); results_outer_frame.grid_columnconfigure(0, weight=1)

        # Protocol handler for closing the window
        master.protocol("WM_DELETE_WINDOW", self.on_closing)

    # --- All Methods (_configure_styles, set_status, validate_input, clear_fields,
    # --- handle_key_event, run_interpretation, format_results_text, copy_summary_to_clipboard,
    # --- export_to_pdf, export_to_excel, change_theme, show_flowchart, show_about,
    # --- show_history_window, show_batch_format_help, import_batch, _parse_header,
    # --- _process_row_data, _process_csv, _process_excel, show_batch_results_window,
    # --- export_batch_pdf, export_batch_excel, show_worklist_window, _worklist_add,
    # --- _worklist_remove, _worklist_load, show_log_viewer_window, run_self_test,
    # --- update_dashboard, show_report_window, _generate_report, _write_summary_pdf,
    # --- _write_summary_excel, on_closing) are included below ---

    def _configure_styles(self):
        """Configure custom styles."""
        self.style.configure('TLabel', font=('Segoe UI', 10)); self.style.configure('TButton', font=('Segoe UI', 10, 'bold'), padding=5)
        self.style.configure('TEntry', font=('Segoe UI', 10), padding=3); self.style.configure('Header.TLabel', font=('Segoe UI', 12, 'bold'))
        self.style.configure('ResultPOS.TLabel', font=('Segoe UI', 16, 'bold'), foreground='red'); self.style.configure('ResultNEG.TLabel', font=('Segoe UI', 16, 'bold'), foreground='darkgreen')
        self.style.configure('ResultIND.TLabel', font=('Segoe UI', 16, 'bold'), foreground='red'); self.style.configure('ResultDefault.TLabel', font=('Segoe UI', 16, 'bold'), foreground='black')
        self.style.configure('Status.TLabel', font=('Segoe UI', 9), padding=2)

    def set_status(self, message):
        self.status_var.set(message); self.master.update_idletasks()

    def validate_input(self, value_str):
        if not value_str: return False, "Input cannot be empty."
        try: return True, float(value_str)
        except ValueError: return False, "Invalid input. Please enter numeric values."

    def clear_fields(self):
        self.sample_id_entry.delete(0, tk.END); self.nil_entry.delete(0, tk.END); self.tb1_entry.delete(0, tk.END); self.tb2_entry.delete(0, tk.END); self.mit_entry.delete(0, tk.END)
        self.results_text.config(state='normal'); self.results_text.delete('1.0', tk.END); self.results_text.config(state='disabled')
        self.result_label_var.set("---"); self.result_label.configure(style='ResultDefault.TLabel'); self.export_pdf_button.config(state=tk.DISABLED); self.export_excel_button.config(state=tk.DISABLED); self.copy_button.config(state=tk.DISABLED); self.last_results = None; self.sample_id_entry.focus_set(); self.set_status("Fields Cleared. Ready.")

    def handle_key_event(self, event):
        """Handles key presses for potential barcode scanning."""
        focused_widget = self.master.focus_get()
        if focused_widget not in (self.sample_id_entry, self.run_id_entry): self._barcode_buffer = ""; return
        current_time = time.time()
        if (current_time - self._last_key_time) < BARCODE_INTERKEY_DELAY:
            if event.char and event.char.isprintable(): self._barcode_buffer += event.char
        else:
            if event.char and event.char.isprintable(): self._barcode_buffer = event.char
            else: self._barcode_buffer = ""
        self._last_key_time = current_time
        if event.keysym == 'Return' and len(self._barcode_buffer) >= BARCODE_MIN_LEN:
            current_value = self._barcode_buffer; focused_widget.delete(0, tk.END); focused_widget.insert(0, current_value); self._barcode_buffer = ""
            if focused_widget == self.sample_id_entry: self.nil_entry.focus_set()
            elif focused_widget == self.run_id_entry: self.sample_id_entry.focus_set()
            return "break"

    def run_interpretation(self, event=None, input_data=None):
        """Interprets a SINGLE sample, displays, logs, saves, performs Delta Check. Returns results_dict or None."""
        if self.is_processing: print("DEBUG: Interpretation already in progress. Skipping."); return None
        self.is_processing = True; original_interpret_state = tk.NORMAL; original_import_state = tk.NORMAL
        if input_data is None:
            try: original_interpret_state=self.interpret_button.cget('state'); original_import_state=self.import_button.cget('state'); self.interpret_button.config(state=tk.DISABLED); self.import_button.config(state=tk.DISABLED); self.master.update_idletasks()
            except tk.TclError: pass
        run_id = ""; sample_id = ""; op_id = ""; warnings = []; results_dict = None
        try:
            if input_data is None:
                self.set_status("Processing Single Sample..."); op_id=self.op_id_entry.get().strip() or "N/A"; run_id=self.run_id_entry.get().strip() or "N/A"; sample_id=self.sample_id_entry.get().strip()
                if not sample_id: messagebox.showerror("Input Error", "Sample ID required."); self.sample_id_entry.focus_set(); self.set_status("Error: Sample ID required."); return None
                inputs = {"Nil": self.nil_entry.get(), "TB1": self.tb1_entry.get(), "TB2": self.tb2_entry.get(), "Mitogen": self.mit_entry.get()}; validated_values = {}; error_messages = []
                for name, value_str in inputs.items():
                    is_valid, result = self.validate_input(value_str)
                    if not is_valid: error_messages.append(f"{name}: {result}")
                    else: validated_values[name] = result
                if error_messages: messagebox.showerror("Input Error", "\n".join(error_messages)); self.last_results = None; self.set_status("Error: Invalid input."); return None
                nil_val=validated_values["Nil"]; tb1_val=validated_values["TB1"]; tb2_val=validated_values["TB2"]; mit_val=validated_values["Mitogen"]
            else: op_id=input_data['operator_id']; run_id=input_data['run_id']; sample_id=input_data['sample_id']; nil_val=input_data['nil']; tb1_val=input_data['tb1']; tb2_val=input_data['tb2']; mit_val=input_data['mitogen']

            previous_db_result = get_previous_result(sample_id)
            if input_data is None:
                if 1.0 < nil_val <= 8.0: warnings.append(f"Nil ({nil_val:.3f}) high but acceptable.")
                if mit_val > 15.0: warnings.append(f"Mitogen ({mit_val:.3f}) very high.")

            try:
                 results_dict = interpret_qft(nil_val, tb1_val, tb2_val, mit_val); results_dict["sample_id"]=sample_id; results_dict["operator_id"]=op_id; results_dict["run_id"]=run_id; results_dict["input_nil"]=nil_val; results_dict["input_tb1"]=tb1_val; results_dict["input_tb2"]=tb2_val; results_dict["input_mit"]=mit_val;
            except Exception as e:
                 current_sample_id_for_error = sample_id if sample_id else "N/A"; error_msg = f"Calculation Error for {current_sample_id_for_error}:\n{e}"; print(error_msg, file=sys.stderr); log_event("ERROR", sample_id=current_sample_id_for_error, details=f"Interpretation calculation failed: {e}\n{traceback.format_exc()}")
                 if input_data is None: messagebox.showerror("Calculation Error", error_msg); self.set_status("Error: Calculation failed.")
                 self.last_results = None; return None

            if previous_db_result:
                prev_res_val, prev_ts = previous_db_result; curr_res_val = results_dict['result']
                if check_significant_change(prev_res_val, curr_res_val):
                    try: prev_dt = datetime.strptime(prev_ts, '%Y-%m-%d %H:%M:%S'); prev_ts_nice = prev_dt.strftime('%Y-%m-%d %H:%M')
                    except: prev_ts_nice = prev_ts
                    warnings.append(f"DELTA CHECK: Result changed significantly from '{prev_res_val}' ({prev_ts_nice}) to '{curr_res_val}'.")

            if warnings and input_data is None: messagebox.showwarning("Interpretation Warnings", "Please note:\n\n" + "\n".join(warnings))

            log_details_subset = {'nil': f"{results_dict['input_nil']:.3f}", 'tb1': f"{results_dict['input_tb1']:.3f}", 'tb2': f"{results_dict['input_tb2']:.3f}", 'mit': f"{results_dict['input_mit']:.3f}" }
            log_event(event_type="INTERPRET", op_id=results_dict['operator_id'], run_id=results_dict['run_id'], sample_id=results_dict['sample_id'], result=results_dict['result'], reason=results_dict['reason'], details=log_details_subset)

            save_successful = save_interpretation_to_db(results_dict)
            if not save_successful and input_data is None: self.set_status("Warning: Failed to save to history DB.")

            if input_data is None:
                self.last_results = results_dict; output_string = self.format_results_text(results_dict); self.results_text.config(state='normal'); self.results_text.delete('1.0', tk.END); self.results_text.insert(tk.END, output_string)
                final_result = results_dict['result']; self.result_label_var.set(final_result)
                if "POS" in final_result: self.result_label.configure(style='ResultPOS.TLabel')
                elif "NEG" in final_result: self.result_label.configure(style='ResultNEG.TLabel')
                elif "IND" in final_result: self.result_label.configure(style='ResultIND.TLabel')
                else: self.result_label.configure(style='ResultDefault.TLabel')
                result_line_index = None;
                for i, line in enumerate(output_string.splitlines()):
                     if line.strip().startswith("QFT RESULT"): result_line_index = i + 1; break
                if result_line_index:
                     line_start = f"{result_line_index}.0"; line_end = f"{result_line_index}.end"; self.results_text.tag_remove("pos_ind_result", "1.0", tk.END); self.results_text.tag_remove("neg_result", "1.0", tk.END)
                     if "IND" in results_dict["result"] or "POS" in results_dict["result"]: self.results_text.tag_add("pos_ind_result", line_start, line_end)
                     elif "NEG" in results_dict["result"]: self.results_text.tag_add("neg_result", line_start, line_end)
                self.results_text.config(state='disabled');
                if REPORTLAB_AVAILABLE: self.export_pdf_button.config(state=tk.NORMAL)
                if OPENPYXL_AVAILABLE: self.export_excel_button.config(state=tk.NORMAL)
                self.copy_button.config(state=tk.NORMAL)
                status_msg = "Interpretation Complete. Result saved and logged."
                if warnings: status_msg += " (Warnings noted)"
                self.set_status(status_msg)
            return results_dict
        finally:
            self.is_processing = False
            if input_data is None:
                try: self.interpret_button.config(state=original_interpret_state); self.import_button.config(state=original_import_state)
                except (tk.TclError, NameError): pass

    def format_results_text(self, r):
        """Formats the results dictionary for the main text widget."""
        header = "="*70+"\n"; header += f" Sample ID: {r.get('sample_id', 'N/A')} | Run ID: {r.get('run_id', 'N/A')} | Operator ID: {r.get('operator_id', 'N/A')}\n"; header += "-"*70+"\n\n"
        table = f"{'Parameter':<15} | {'Input (IU/mL)':<15} | {'Calculated Value':<30}\n"; table += "-"*70+"\n"
        table += f"{'Nil':<15} | {r.get('input_nil', 0.0):<15.3f} | {'25% of Nil:':<14} {r.get('nil_25', 0.0):<15.4f}\n"; table += f"{'TB1':<15} | {r.get('input_tb1', 0.0):<15.3f} | {'TB1 - Nil:':<14} {r.get('tb1_nil', 0.0):<15.4f}\n"
        table += f"{'TB2':<15} | {r.get('input_tb2', 0.0):<15.3f} | {'TB2 - Nil:':<14} {r.get('tb2_nil', 0.0):<15.4f}\n"; table += f"{'Mitogen':<15} | {r.get('input_mit', 0.0):<15.3f} | {'Mit - Nil:':<14} {r.get('mit_nil', 0.0):<15.4f}\n"; table += "-"*70+"\n"
        table += f"{'QFT RESULT':<15} | {r.get('result', 'Error'):<53}\n"
        if r.get('reason'): table += f"{'Reason':<15} | {r.get('reason', ''):<53}\n"
        table += "="*70+"\n"; table += "*IND: Indeterminate; †POS: Positive; NEG: Negative\n"; table += "Note: Refer to Flowchart (Help Menu). Clinical correlation required.\n"
        return header + table

    def copy_summary_to_clipboard(self):
        """Copies the content of the results text area to the clipboard."""
        if not self.last_results: self.set_status("Nothing to copy."); return
        try: summary_text = self.results_text.get("1.0", tk.END).strip(); self.master.clipboard_clear(); self.master.clipboard_append(summary_text); self.set_status("Summary copied to clipboard.")
        except Exception as e: messagebox.showerror("Copy Error", f"Could not copy text:\n{e}"); self.set_status("Error copying summary.")

    def export_to_pdf(self):
        """Exports the CURRENT SINGLE result to a PDF file."""
        if not REPORTLAB_AVAILABLE: messagebox.showerror("PDF Export Error", "ReportLab library not installed."); return
        if not self.last_results: messagebox.showwarning("Export Error", "No results available for current sample."); return
        r=self.last_results; timestamp=datetime.now().strftime("%Y%m%d_%H%M%S"); safe_sample_id="".join(c if c.isalnum() else"_" for c in r['sample_id']); suggested_filename=f"QFT_Result_{safe_sample_id}_{timestamp}.pdf"
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Documents", "*.pdf")], initialfile=suggested_filename, title="Save QFT Result as PDF")
        if not filepath: self.set_status("PDF Export Cancelled."); return
        self.set_status("Exporting PDF...")
        try:
            doc = SimpleDocTemplate(filepath, pagesize=(8.5*inch, 11*inch), leftMargin=0.75*inch, rightMargin=0.75*inch, topMargin=0.75*inch, bottomMargin=0.75*inch); styles = getSampleStyleSheet(); story = []
            story.append(Paragraph("LIAISON® QuantiFERON-TB® Gold Plus Interpretation Report", styles['h1'])); story.append(Spacer(1, 0.1*inch)); story.append(Paragraph(f"<b>Sample ID:</b> {r['sample_id']}", styles['h3'])); story.append(Paragraph(f"<b>Run ID:</b> {r.get('run_id', 'N/A')}", styles['Normal'])); story.append(Paragraph(f"<b>Operator ID:</b> {r['operator_id']}", styles['Normal'])); story.append(Spacer(1, 0.1*inch)); story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])); story.append(Spacer(1, 0.2*inch))
            data=[['Parameter','Input (IU/mL)','Calculated Value',''],['Nil',f"{r['input_nil']:.3f}",'25% of Nil:',f"{r['nil_25']:.4f}"],['TB Antigen 1 (TB1)',f"{r['input_tb1']:.3f}",'TB1 - Nil:',f"{r['tb1_nil']:.4f}"],['TB Antigen 2 (TB2)',f"{r['input_tb2']:.3f}",'TB2 - Nil:',f"{r['tb2_nil']:.4f}"],['Mitogen (Mit)',f"{r['input_mit']:.3f}",'Mit - Nil:',f"{r['mit_nil']:.4f}"]]; table=Table(data, colWidths=[1.5*inch,1.5*inch,1.5*inch,1.5*inch]); table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('BOTTOMPADDING',(0,0),(-1,0),12),('BACKGROUND',(0,1),(-1,-1),colors.beige),('GRID',(0,0),(-1,-1),1,colors.black),('FONTSIZE',(0,0),(-1,-1),10),('ALIGN',(1,1),(1,-1),'RIGHT'),('ALIGN',(3,1),(3,-1),'RIGHT'),('RIGHTPADDING',(1,1),(1,-1),10),('RIGHTPADDING',(3,1),(3,-1),10)])); story.append(table); story.append(Spacer(1, 0.3*inch))
            res_style=styles['h2'];
            if "IND" in r['result'] or "POS" in r['result']: res_style.textColor = colors.red
            elif "NEG" in r['result']: res_style.textColor = colors.darkgreen
            else: res_style.textColor = colors.black
            story.append(Paragraph(f"QFT RESULT: {r['result']}", res_style)); story.append(Spacer(1, 0.1*inch)); story.append(Paragraph(f"Reason: {r['reason']}", styles['Normal'])); story.append(Spacer(1, 0.3*inch))
            disclaimer_style=styles['Italic']; disclaimer_style.fontSize=9; story.append(Paragraph("Disclaimer: This report was generated using an automated tool based on the manufacturer's algorithm (Figure 1 - Viewable via Help Menu). Results should always be interpreted in the context of the patient's clinical information, risk factors, and other diagnostic findings. This tool does not replace professional medical judgment.", disclaimer_style))
            doc.build(story); messagebox.showinfo("Export Successful", f"Results exported to:\n{filepath}"); self.set_status("PDF Export Successful.")
        except PermissionError: messagebox.showerror("Export Error", f"Permission denied: {filepath}"); self.set_status("Error: PDF Permission Denied.")
        except Exception as e: messagebox.showerror("PDF Export Error", f"Error creating PDF:\n{e}"); log_event("ERROR", details=f"PDF Export failed: {e}\n{traceback.format_exc()}"); self.set_status("Error: PDF Export Failed.")

    def export_to_excel(self):
        """Exports the CURRENT SINGLE result to an Excel file."""
        if not OPENPYXL_AVAILABLE: messagebox.showerror("Excel Export Error", "openpyxl library not installed."); return
        if not self.last_results: messagebox.showwarning("Export Error", "No results available for current sample."); return
        r = self.last_results; timestamp = datetime.now().strftime("%Y%m%d_%H%M%S"); safe_sample_id="".join(c if c.isalnum() else"_" for c in r['sample_id']); suggested_filename=f"QFT_Result_{safe_sample_id}_{timestamp}.xlsx"
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")], initialfile=suggested_filename, title="Save QFT Result as Excel")
        if not filepath: self.set_status("Excel Export Cancelled."); return
        self.set_status("Exporting Excel...")
        try:
            wb = Workbook(); ws = wb.active; ws.title = "QFT Interpretation"
            header_font=Font(bold=True,size=12); title_font=Font(bold=True,size=14); bold_font=Font(bold=True); center_align=Alignment(horizontal='center',vertical='center'); right_align=Alignment(horizontal='right',vertical='center'); left_align=Alignment(horizontal='left',vertical='top'); wrap_align=Alignment(wrap_text=True,vertical='top'); thin_border_side=Side(border_style="thin",color="000000"); thin_border=Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side); red_fill=PatternFill(start_color="FFFFC7CE",end_color="FFFFC7CE",fill_type="solid"); green_fill=PatternFill(start_color="FFC6EFCE",end_color="FFC6EFCE",fill_type="solid")
            ws.merge_cells('A1:D1'); ws['A1'] = "LIAISON® QuantiFERON-TB® Gold Plus Interpretation Report"; ws['A1'].font = title_font; ws['A1'].alignment = center_align
            ws['A2'] = "Sample ID:"; ws['A2'].font = bold_font; ws.merge_cells('B2:D2'); ws['B2'] = r['sample_id']; ws['B2'].font = bold_font
            ws['A3'] = "Run ID:"; ws['A3'].font = bold_font; ws.merge_cells('B3:D3'); ws['B3'] = r.get('run_id', 'N/A')
            ws['A4'] = "Operator ID:"; ws['A4'].font = bold_font; ws.merge_cells('B4:D4'); ws['B4'] = r['operator_id']
            ws['A5'] = "Report Generated:"; ws['A5'].font = bold_font; ws.merge_cells('B5:D5'); ws['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            current_row = 7; headers=["Parameter","Input (IU/mL)","Calculated Value","Value"];
            for col_idx, header in enumerate(headers, 1): ws.cell(row=current_row, column=col_idx, value=header).font = header_font
            current_row += 1; data_rows=[("Nil",r['input_nil'],"25% of Nil:",r['nil_25']),("TB Antigen 1 (TB1)",r['input_tb1'],"TB1 - Nil:",r['tb1_nil']),("TB Antigen 2 (TB2)",r['input_tb2'],"TB2 - Nil:",r['tb2_nil']),("Mitogen (Mit)",r['input_mit'],"Mit - Nil:",r['mit_nil'])]
            for i, row_data in enumerate(data_rows): ws.cell(row=current_row+i, column=1, value=row_data[0]); ws.cell(row=current_row+i, column=2, value=row_data[1]).number_format='0.000'; ws.cell(row=current_row+i, column=3, value=row_data[2]); ws.cell(row=current_row+i, column=4, value=row_data[3]).number_format='0.0000'
            data_end_row = current_row + len(data_rows) - 1
            for row_idx in range(7, data_end_row + 1):
                for col_idx in range(1, 5): cell=ws.cell(row=row_idx, column=col_idx); cell.border=thin_border;
                if col_idx == 1 or col_idx == 3: cell.alignment = left_align
                else: cell.alignment = right_align
            current_row = data_end_row + 2
            ws.cell(row=current_row, column=1, value="QFT RESULT:").font=header_font; ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4); result_cell=ws.cell(row=current_row, column=2, value=r['result']); result_cell.font=bold_font; result_cell.alignment=center_align;
            if "IND" in r['result'] or "POS" in r['result']: result_cell.fill = red_fill
            elif "NEG" in r['result']: result_cell.fill = green_fill
            current_row += 1; ws.cell(row=current_row, column=1, value="Reason:").font=header_font; ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4); reason_cell=ws.cell(row=current_row, column=2, value=r['reason']); reason_cell.alignment=wrap_align; current_row += 2
            ws.cell(row=current_row, column=1, value="Disclaimer:").font=bold_font; ws.merge_cells(start_row=current_row + 1, start_column=1, end_row=current_row + 3, end_column=4); disclaimer_cell=ws.cell(row=current_row + 1, column=1, value="Disclaimer: This report was generated using an automated tool based on the manufacturer's algorithm (Figure 1 - Viewable via Help Menu). Results should always be interpreted in the context of the patient's clinical information, risk factors, and other diagnostic findings. This tool does not replace professional medical judgment."); disclaimer_cell.font=Font(italic=True, size=9); disclaimer_cell.alignment=wrap_align
            ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=18; ws.column_dimensions['C'].width=18; ws.column_dimensions['D'].width=18
            wb.save(filepath); messagebox.showinfo("Export Successful", f"Results exported to:\n{filepath}"); self.set_status("Excel Export Successful.")
        except PermissionError: messagebox.showerror("Export Error", f"Permission denied: {filepath}"); self.set_status("Error: Excel Permission Denied.")
        except Exception as e: messagebox.showerror("Excel Export Error", f"Error creating Excel:\n{e}"); log_event("ERROR", details=f"Excel Export failed: {e}\n{traceback.format_exc()}"); self.set_status("Error: Excel Export Failed.")

    # --- Menu Commands ---
    def change_theme(self):
        """Applies the selected theme."""
        selected_theme = self.current_theme.get()
        try: self.style.theme_use(selected_theme); self._configure_styles(); self.config['theme'] = selected_theme; self.set_status(f"Theme changed to '{selected_theme}'.")
        except tk.TclError: messagebox.showerror("Theme Error", f"Could not apply theme '{selected_theme}'."); self.current_theme.set(self.style.theme_use())

    def show_flowchart(self):
        """Displays the flowchart image."""
        if not PIL_AVAILABLE: messagebox.showerror("Error", "Pillow (PIL) library required."); return
        if not os.path.exists(FLOWCHART_PATH): messagebox.showerror("Error", f"Flowchart image not found:\n{FLOWCHART_PATH}"); return
        try:
            flowchart_window=tk.Toplevel(self.master); flowchart_window.title("Interpretation Flowchart (Figure 1)"); flowchart_window.transient(self.master)
            img=Image.open(FLOWCHART_PATH); flowchart_window.image_tk=ImageTk.PhotoImage(img); img_label=tk.Label(flowchart_window,image=flowchart_window.image_tk); img_label.pack(padx=10, pady=10)
            flowchart_window.grab_set(); flowchart_window.focus_set(); flowchart_window.wait_window()
        except Exception as e: messagebox.showerror("Image Error", f"Failed to display flowchart:\n{e}"); log_event("ERROR", details=f"Flowchart display failed: {e}\n{traceback.format_exc()}")

    def show_about(self):
        """Displays the About dialog box with System Info."""
        try: python_version = platform.python_version()
        except Exception: python_version = "N/A"
        try: os_version = platform.platform()
        except Exception: os_version = "N/A"
        try: tk_version = tk.TkVersion
        except Exception: tk_version = "N/A"
        try: from PIL import __version__ as pil_version
        except ImportError: pil_version = "Not Installed"
        try: from openpyxl import __version__ as openpyxl_version
        except ImportError: openpyxl_version = "Not Installed"
        try: from reportlab import Version as rl_version
        except ImportError: rl_version = "Not Installed"
        sqlite_version = sqlite3.sqlite_version
        info_text = f"""{APP_TITLE}\n\nInterpreter based on LIAISON® QuantiFERON-TB® Gold Plus Assay.\nFor Research Use Only or Laboratory Use where applicable.\nDisclaimer: Always correlate results with clinical findings.\n\n--- System Information ---\nPython Version: {python_version}\nOperating System: {os_version}\nTk Version: {tk_version}\nPillow Version: {pil_version}\nOpenpyxl Version: {openpyxl_version}\nReportLab Version: {rl_version}\nSQLite Version: {sqlite_version}\n\n--- File Paths ---\nScript Directory: {SCRIPT_DIR}\nConfig File: {CONFIG_FILENAME}\nLog File (CSV): {LOG_FILENAME}\nHistory DB: {DB_FILENAME}"""
        messagebox.showinfo(f"About {APP_TITLE}", info_text)

    def show_history_window(self):
        """Creates and shows the history browsing window."""
        history_win = tk.Toplevel(self.master); history_win.title("Interpretation History"); history_win.geometry("1050x600"); history_win.transient(self.master)
        controls_frame = ttk.Frame(history_win, padding="10"); controls_frame.pack(fill=tk.X, side=tk.TOP)
        ttk.Label(controls_frame, text="Sample ID:").grid(row=0, column=0, padx=(0,2),pady=5,sticky=tk.W); sample_search_entry = ttk.Entry(controls_frame, width=15); sample_search_entry.grid(row=0, column=1, padx=(0,10),pady=5,sticky=tk.W)
        ttk.Label(controls_frame, text="Run ID:").grid(row=0, column=2, padx=(0,2),pady=5,sticky=tk.W); run_search_entry = ttk.Entry(controls_frame, width=15); run_search_entry.grid(row=0, column=3, padx=(0,10),pady=5,sticky=tk.W)
        ttk.Label(controls_frame, text="Date (YYYY-MM-DD):").grid(row=0, column=4, padx=(0,2),pady=5,sticky=tk.W); date_search_entry = ttk.Entry(controls_frame, width=12); date_search_entry.grid(row=0, column=5, padx=(0,10),pady=5,sticky=tk.W)
        search_button = ttk.Button(controls_frame, text="Search / Filter", command=lambda: load_history(tv, sample_search_entry.get(), date_search_entry.get(), run_search_entry.get())); search_button.grid(row=0, column=6, padx=5, pady=5)
        show_all_button = ttk.Button(controls_frame, text="Show All (Recent)", command=lambda: load_history(tv, None, None, None)); show_all_button.grid(row=0, column=7, padx=5, pady=5)
        close_button = ttk.Button(controls_frame, text="Close", command=history_win.destroy); close_button.grid(row=0, column=8, padx=(20,5), pady=5)
        tree_frame = ttk.Frame(history_win, padding="10"); tree_frame.pack(fill=tk.BOTH, expand=True)
        columns = ("timestamp", "operator", "run_id", "sample_id", "nil", "tb1", "tb2", "mit", "result", "reason"); tv = ttk.Treeview(tree_frame, columns=columns, show='headings', height=20)
        tv.heading("timestamp", text="Timestamp", anchor=tk.W); tv.heading("operator", text="Operator", anchor=tk.W); tv.heading("run_id", text="Run ID", anchor=tk.W); tv.heading("sample_id", text="Sample ID", anchor=tk.W); tv.heading("nil", text="Nil", anchor=tk.E); tv.heading("tb1", text="TB1", anchor=tk.E); tv.heading("tb2", text="TB2", anchor=tk.E); tv.heading("mit", text="Mitogen", anchor=tk.E); tv.heading("result", text="Result", anchor=tk.W); tv.heading("reason", text="Reason", anchor=tk.W)
        tv.column("timestamp", width=140, stretch=tk.NO, anchor=tk.W); tv.column("operator", width=70, stretch=tk.NO, anchor=tk.W); tv.column("run_id", width=100, stretch=tk.NO, anchor=tk.W); tv.column("sample_id", width=120, stretch=tk.NO, anchor=tk.W); tv.column("nil", width=60, stretch=tk.NO, anchor=tk.E); tv.column("tb1", width=60, stretch=tk.NO, anchor=tk.E); tv.column("tb2", width=60, stretch=tk.NO, anchor=tk.E); tv.column("mit", width=60, stretch=tk.NO, anchor=tk.E); tv.column("result", width=70, stretch=tk.NO, anchor=tk.W); tv.column("reason", width=250, stretch=tk.YES, anchor=tk.W)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tv.yview); hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tv.xview); tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tv.grid(row=0, column=0, sticky='nsew'); vsb.grid(row=0, column=1, sticky='ns'); hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1); tree_frame.grid_columnconfigure(0, weight=1)
        load_history(tv, None, None, None); history_win.grab_set(); history_win.focus_set(); history_win.wait_window()

        # --- Help/Guide Methods ---
        # --- Help/Guide Methods ---
    def show_batch_format_help(self):
        """Displays a Toplevel window explaining the batch import format using grid."""
        help_win = tk.Toplevel(self.master)
        help_win.title("Batch Import Format Guide")
        # help_win.geometry("550x400") # Remove fixed geometry
        help_win.minsize(550, 420) # Set a minimum practical size
        help_win.resizable(False, True) # Allow vertical resizing

        # Ensure this window stays on top of the main app while open
        help_win.transient(self.master)

        # Use a main frame that fills the window and use grid within it
        main_frame = ttk.Frame(help_win, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH) # Frame fills the window

        # Configure grid columns for the frame
        main_frame.columnconfigure(0, weight=1) # Make column 0 expandable

        # --- Widgets using grid ---
        row_index = 0

        # Title
        title_label = ttk.Label(main_frame, text="Batch Import File Format Requirements", style='Header.TLabel')
        title_label.grid(row=row_index, column=0, sticky="w", pady=(0, 10))
        row_index += 1

        # Instructions using Labels
        instructions = [
            "Supported formats: CSV (.csv) or Excel (.xlsx - reads first sheet).",
            "The first row MUST be a header row containing the required column names.",
            "Required header names (case-insensitive):",
            "   • Sample ID",
            "   • Nil",
            "   • TB1",
            "   • TB2",
            "   • Mitogen",
            "Column order does not matter.",
            "Data should start from the second row.",
            "Numeric columns (Nil, TB1, TB2, Mitogen) should contain only numbers.",
            "Rows with missing Sample ID or invalid numbers will be skipped."
        ]
        for instruction in instructions:
            instr_label = ttk.Label(main_frame, text=instruction, wraplength=500, justify=tk.LEFT) # Wraplength helps control width
            instr_label.grid(row=row_index, column=0, sticky="w", pady=1)
            row_index += 1

        # Example Section Label
        example_label = ttk.Label(main_frame, text="Example Structure:", font=('Segoe UI', 10, 'bold'))
        example_label.grid(row=row_index, column=0, sticky="w", pady=(15, 2))
        row_index += 1

        # Example Text Area
        example_text_widget = tk.Text(main_frame, height=7, width=60, wrap=tk.NONE, # Increased height
                                      font=('Consolas', 9), relief=tk.SOLID, borderwidth=1)
        example_content = """Sample ID,Nil,TB1,TB2,Mitogen
Sample-001,0.120,5.600,4.800,8.900
Sample-002,0.080,0.150,0.200,6.500
Sample-003,0.210,0.450,0.330,7.100
Control-A,9.200,10.00,11.00,15.00"""
        example_text_widget.insert(tk.END, example_content)
        example_text_widget.config(state='disabled') # Make read-only
        example_text_widget.grid(row=row_index, column=0, sticky="ew", pady=5) # Sticky ew makes it expand horizontally
        row_index += 1

        # Configure the row containing the text widget to expand if window is resized vertically
        main_frame.rowconfigure(row_index - 1, weight=1) # Give weight to the text widget's row

        # OK Button Frame (to center the button)
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=row_index, column=0, pady=(15, 0))
        # Configure the button frame's column 0 to center the button
        button_frame.columnconfigure(0, weight=1)

        ok_button = ttk.Button(button_frame, text="OK", command=help_win.destroy, style='TButton')
        ok_button.grid(row=0, column=0) # Place button inside its frame
        ok_button.focus_set()
        row_index += 1


        # Make modal
        help_win.grab_set()
        self.master.wait_window(help_win)
    
    # --- Batch Import Methods ---
    def import_batch(self):
        """Handles batch import."""
        self.show_batch_format_help()
        filepath = filedialog.askopenfilename(title="Select Batch File", filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")])
        if not filepath: self.set_status("Batch import cancelled."); return
        self.set_status(f"Importing batch from {os.path.basename(filepath)}...")
        file_extension = os.path.splitext(filepath)[1].lower(); op_id = self.op_id_entry.get().strip() or "N/A"; run_id = self.run_id_entry.get().strip() or "N/A"
        processed_results = []; skipped_rows = 0; total_rows = 0
        try:
            if file_extension == ".csv": processed_results, skipped_rows, total_rows = self._process_csv(filepath, op_id, run_id)
            elif file_extension == ".xlsx":
                if not OPENPYXL_AVAILABLE: messagebox.showerror("Import Error", "Openpyxl library required for .xlsx files."); self.set_status("Error: Missing openpyxl."); return
                processed_results, skipped_rows, total_rows = self._process_excel(filepath, op_id, run_id)
            else: messagebox.showerror("Import Error", f"Unsupported file type: {file_extension}"); self.set_status("Error: Unsupported file type."); return
            if processed_results or skipped_rows > 0: self.show_batch_results_window(processed_results, skipped_rows, total_rows, os.path.basename(filepath)); self.set_status(f"Batch Import Finished: {len(processed_results)} processed, {skipped_rows} skipped.")
            else: messagebox.showinfo("Batch Import", "No valid data rows found or processed."); self.set_status("Batch Import Finished: No valid data.")
        except Exception as e: messagebox.showerror("Batch Import Error", f"Error during batch processing:\n{e}"); self.set_status("Error during batch import."); log_event("ERROR", details=f"Batch Import failed: {e}\n{traceback.format_exc()}")

    def _parse_header(self, header_row):
        """Parses batch header."""
        header_map = {}; found_headers = set(); lowered_headers = [str(h).strip().lower() if h is not None else '' for h in header_row]
        for i, header in enumerate(lowered_headers):
            if header == 'sample id': header_map['sample_id'] = i; found_headers.add('sample id')
            elif header == 'nil': header_map['nil'] = i; found_headers.add('nil')
            elif header == 'tb1': header_map['tb1'] = i; found_headers.add('tb1')
            elif header == 'tb2': header_map['tb2'] = i; found_headers.add('tb2')
            elif header == 'mitogen': header_map['mitogen'] = i; found_headers.add('mitogen')
        missing_headers = REQUIRED_BATCH_HEADERS - found_headers
        if missing_headers: messagebox.showerror("Header Error", f"Missing headers:\n{', '.join(sorted(list(missing_headers)))}"); return None
        return header_map

    def _process_row_data(self, row_values, header_map, op_id, run_id, row_num):
        """Processes single batch row."""
        try:
            sample_id = str(row_values[header_map['sample_id']]).strip(); nil_str = str(row_values[header_map['nil']]).strip(); tb1_str = str(row_values[header_map['tb1']]).strip(); tb2_str = str(row_values[header_map['tb2']]).strip(); mit_str = str(row_values[header_map['mitogen']]).strip()
            if not sample_id: print(f"Skipping row {row_num}: Missing Sample ID.", file=sys.stderr); return None
            try: nil_val = float(nil_str); tb1_val = float(tb1_str); tb2_val = float(tb2_str); mit_val = float(mit_str)
            except (ValueError, TypeError): print(f"Skipping row {row_num} (Sample: {sample_id}): Invalid numeric data.", file=sys.stderr); return None
            input_data = {'operator_id': op_id, 'run_id': run_id, 'sample_id': sample_id, 'nil': nil_val, 'tb1': tb1_val, 'tb2': tb2_val, 'mitogen': mit_val}
            result_dict = self.run_interpretation(input_data=input_data); return result_dict
        except IndexError: print(f"Skipping row {row_num}: Too few columns.", file=sys.stderr); return None
        except Exception as e: print(f"Error processing row {row_num}: {e}", file=sys.stderr); log_event("ERROR", sample_id=sample_id if 'sample_id' in locals() else 'N/A', details=f"Batch row processing error: {e}\n{traceback.format_exc()}"); return None

    def _process_csv(self, filepath, op_id, run_id):
        """Processes CSV."""
        processed_results = []; skipped = 0; row_num = 0; header_map = None
        with open(filepath, 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                row_num += 1
                if row_num == 1: header_map = self._parse_header(row);
                if header_map is None and row_num == 1: return [], 0, 0
                elif header_map and row_num > 1:
                    if not any(row): continue
                    result = self._process_row_data(row, header_map, op_id, run_id, row_num)
                    if result: processed_results.append(result)
                    else: skipped += 1
        return processed_results, skipped, max(0, row_num -1)

    def _process_excel(self, filepath, op_id, run_id):
        """Processes Excel."""
        processed_results = []; skipped = 0; row_num = 0; header_map = None; wb = None
        try:
            wb = load_workbook(filename=filepath, read_only=True, data_only=True); ws = wb.active
            for row in ws.iter_rows():
                row_num += 1; row_values = [cell.value if cell.value is not None else '' for cell in row]
                if row_num == 1: header_map = self._parse_header(row_values);
                if header_map is None and row_num == 1: return [], 0, 0
                elif header_map and row_num > 1:
                     if not any(str(v).strip() for v in row_values): continue
                     result = self._process_row_data(row_values, header_map, op_id, run_id, row_num)
                     if result: processed_results.append(result)
                     else: skipped += 1
        finally:
            if wb: wb.close()
        return processed_results, skipped, max(0, row_num -1)

    def show_batch_results_window(self, results_list, skipped_count, total_rows, filename):
        """Displays batch results."""
        batch_win = tk.Toplevel(self.master); batch_win.title(f"Batch Import Results: {filename}"); batch_win.geometry("1000x650"); batch_win.transient(self.master)
        summary_frame = ttk.Frame(batch_win, padding="10"); summary_frame.pack(fill=tk.X, side=tk.TOP); summary_text = f"Processed: {len(results_list)} / {total_rows} rows. Skipped: {skipped_count} rows."; ttk.Label(summary_frame, text=summary_text, font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        pdf_batch_button = ttk.Button(summary_frame, text="Export Batch PDF", command=lambda: self.export_batch_pdf(results_list, filename), state=(tk.NORMAL if REPORTLAB_AVAILABLE and results_list else tk.DISABLED)); pdf_batch_button.pack(side=tk.RIGHT, padx=5)
        excel_batch_button = ttk.Button(summary_frame, text="Export Batch Excel", command=lambda: self.export_batch_excel(results_list, filename), state=(tk.NORMAL if OPENPYXL_AVAILABLE and results_list else tk.DISABLED)); excel_batch_button.pack(side=tk.RIGHT, padx=5)
        close_button = ttk.Button(summary_frame, text="Close Results", command=batch_win.destroy); close_button.pack(side=tk.RIGHT, padx=5)
        tree_frame = ttk.Frame(batch_win, padding="10"); tree_frame.pack(fill=tk.BOTH, expand=True)
        columns = ("sample_id", "nil", "tb1", "tb2", "mit", "result", "reason"); tv = ttk.Treeview(tree_frame, columns=columns, show='headings', height=25)
        tv.heading("sample_id", text="Sample ID", anchor=tk.W); tv.heading("nil", text="Nil", anchor=tk.E); tv.heading("tb1", text="TB1", anchor=tk.E); tv.heading("tb2", text="TB2", anchor=tk.E); tv.heading("mit", text="Mitogen", anchor=tk.E); tv.heading("result", text="Result", anchor=tk.W); tv.heading("reason", text="Reason", anchor=tk.W)
        tv.column("sample_id", width=150, stretch=tk.NO, anchor=tk.W); tv.column("nil", width=70, stretch=tk.NO, anchor=tk.E); tv.column("tb1", width=70, stretch=tk.NO, anchor=tk.E); tv.column("tb2", width=70, stretch=tk.NO, anchor=tk.E); tv.column("mit", width=70, stretch=tk.NO, anchor=tk.E); tv.column("result", width=80, stretch=tk.NO, anchor=tk.W); tv.column("reason", width=400, stretch=tk.YES, anchor=tk.W)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tv.yview); hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tv.xview); tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tv.grid(row=0, column=0, sticky='nsew'); vsb.grid(row=0, column=1, sticky='ns'); hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1); tree_frame.grid_columnconfigure(0, weight=1)
        if not results_list: tv.insert('', tk.END, values=("No results processed.", "", "", "", "", "", ""))
        else:
            for r in results_list: values = (r.get('sample_id',''), f"{r.get('input_nil',0.0):.3f}", f"{r.get('input_tb1',0.0):.3f}", f"{r.get('input_tb2',0.0):.3f}", f"{r.get('input_mit',0.0):.3f}", r.get('result','Error'), r.get('reason','')); tv.insert('', tk.END, values=values)

    # --- Batch Export Methods ---
    def export_batch_pdf(self, results_list, source_filename):
        """Exports a list of batch results to a single PDF file."""
        if not REPORTLAB_AVAILABLE: messagebox.showerror("PDF Export Error", "ReportLab library not installed."); return
        if not results_list: messagebox.showwarning("Export Error", "No batch results available to export."); return
        timestamp=datetime.now().strftime("%Y%m%d_%H%M%S"); suggested_filename=f"QFT_Batch_{os.path.splitext(source_filename)[0]}_{timestamp}.pdf"
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Documents", "*.pdf")], initialfile=suggested_filename, title="Save Batch Results as PDF")
        if not filepath: return
        self.set_status("Exporting Batch PDF...")
        try:
            doc = SimpleDocTemplate(filepath, pagesize=(11*inch, 8.5*inch), leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch); styles = getSampleStyleSheet(); story = []
            story.append(Paragraph("LIAISON® QuantiFERON-TB® Gold Plus - Batch Interpretation Report", styles['h1'])); story.append(Paragraph(f"Source File: {source_filename}", styles['Normal']))
            run_id_batch = results_list[0].get('run_id', 'N/A') if results_list else 'N/A'; story.append(Paragraph(f"Run ID: {run_id_batch}", styles['Normal']))
            story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])); story.append(Spacer(1, 0.2*inch))
            headers = ["Sample ID", "Operator", "Nil", "TB1", "TB2", "Mit", "Result", "Reason"]; col_widths = [1.5*inch, 1.0*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.8*inch, 3.0*inch]
            table_data = [headers]
            for r in results_list: row_data = [r.get('sample_id',''), r.get('operator_id',''), f"{r.get('input_nil',0.0):.3f}", f"{r.get('input_tb1',0.0):.3f}", f"{r.get('input_tb2',0.0):.3f}", f"{r.get('input_mit',0.0):.3f}", r.get('result','Error'), Paragraph(r.get('reason',''), styles['Normal'])]; table_data.append(row_data)
            table = Table(table_data, colWidths=col_widths); table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,0),'CENTER'), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,0),10), ('TOPPADDING',(0,0),(-1,0),4),('BACKGROUND',(0,1),(-1,-1),colors.whitesmoke),('GRID',(0,0),(-1,-1),0.5,colors.black),('ALIGN',(2,1),(6,-1),'RIGHT')]))
            story.append(KeepTogether(table)); story.append(Spacer(1, 0.3*inch))
            disclaimer_style = styles['Italic']; disclaimer_style.fontSize = 9; story.append(Paragraph("Disclaimer: This report was generated using an automated tool based on the manufacturer's algorithm. Results should always be interpreted in the context of the patient's clinical information, risk factors, and other diagnostic findings. This tool does not replace professional medical judgment.", disclaimer_style))
            doc.build(story); messagebox.showinfo("Export Successful", f"Batch results exported to:\n{filepath}"); self.set_status("Batch PDF Export Successful.")
        except PermissionError: messagebox.showerror("Export Error", f"Permission denied: {filepath}"); self.set_status("Error: Batch PDF Permission Denied.")
        except Exception as e: messagebox.showerror("PDF Export Error", f"Error creating batch PDF:\n{e}"); log_event("ERROR", details=f"Batch PDF Export failed: {e}\n{traceback.format_exc()}"); self.set_status("Error: Batch PDF Export Failed.")

    def export_batch_excel(self, results_list, source_filename):
        """Exports a list of batch results to a single Excel file."""
        if not OPENPYXL_AVAILABLE: messagebox.showerror("Excel Export Error", "openpyxl library not installed."); return
        if not results_list: messagebox.showwarning("Export Error", "No batch results available to export."); return
        timestamp=datetime.now().strftime("%Y%m%d_%H%M%S"); suggested_filename=f"QFT_Batch_{os.path.splitext(source_filename)[0]}_{timestamp}.xlsx"
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")], initialfile=suggested_filename, title="Save Batch Results as Excel")
        if not filepath: return
        self.set_status("Exporting Batch Excel...")
        try:
            wb = Workbook(); ws = wb.active; ws.title = "QFT Batch Results"
            header_font=Font(bold=True, size=11); bold_font=Font(bold=True); right_align=Alignment(horizontal='right', vertical='center'); left_align=Alignment(horizontal='left', vertical='top'); wrap_align=Alignment(wrap_text=True, vertical='top'); thin_border_side=Side(border_style="thin", color="000000"); thin_border=Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side); red_fill=PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"); green_fill=PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
            ws['A1'] = "LIAISON® QuantiFERON-TB® Gold Plus - Batch Interpretation Report"; ws['A2'] = "Source File:"; ws['B2'] = source_filename; run_id_batch = results_list[0].get('run_id', 'N/A') if results_list else 'N/A'; ws['A3'] = "Run ID:"; ws['B3'] = run_id_batch; ws['A4'] = "Report Generated:"; ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S'); ws.cell(row=2,column=1).font=bold_font; ws.cell(row=3,column=1).font=bold_font; ws.cell(row=4,column=1).font=bold_font
            current_row = 6; headers=["Sample ID","Operator","Nil","TB1","TB2","Mitogen","Result","Reason","Nil-25%","TB1-Nil","TB2-Nil","Mit-Nil"]
            for col_idx, header in enumerate(headers, 1): cell = ws.cell(row=current_row, column=col_idx, value=header); cell.font = header_font; cell.border = thin_border
            current_row += 1
            for r in results_list:
                ws.cell(row=current_row, column=1, value=r.get('sample_id','')); ws.cell(row=current_row, column=1).alignment=left_align; ws.cell(row=current_row, column=2, value=r.get('operator_id','')); ws.cell(row=current_row, column=2).alignment=left_align
                ws.cell(row=current_row, column=3, value=r.get('input_nil')).number_format='0.000'; ws.cell(row=current_row, column=3).alignment=right_align; ws.cell(row=current_row, column=4, value=r.get('input_tb1')).number_format='0.000'; ws.cell(row=current_row, column=4).alignment=right_align
                ws.cell(row=current_row, column=5, value=r.get('input_tb2')).number_format='0.000'; ws.cell(row=current_row, column=5).alignment=right_align; ws.cell(row=current_row, column=6, value=r.get('input_mit')).number_format='0.000'; ws.cell(row=current_row, column=6).alignment=right_align
                result_cell=ws.cell(row=current_row, column=7, value=r.get('result','Error')); result_cell.alignment=left_align; result_cell.font=bold_font
                reason_cell=ws.cell(row=current_row, column=8, value=r.get('reason','')); reason_cell.alignment=wrap_align
                ws.cell(row=current_row, column=9, value=r.get('nil_25')).number_format='0.0000'; ws.cell(row=current_row, column=9).alignment=right_align; ws.cell(row=current_row, column=10, value=r.get('tb1_nil')).number_format='0.0000'; ws.cell(row=current_row, column=10).alignment=right_align
                ws.cell(row=current_row, column=11, value=r.get('tb2_nil')).number_format='0.0000'; ws.cell(row=current_row, column=11).alignment=right_align; ws.cell(row=current_row, column=12, value=r.get('mit_nil')).number_format='0.0000'; ws.cell(row=current_row, column=12).alignment=right_align
                for col_idx in range(1, len(headers)+1): ws.cell(row=current_row, column=col_idx).border=thin_border
                if "IND" in r['result'] or "POS" in r['result']: result_cell.fill=red_fill
                elif "NEG" in r['result']: result_cell.fill=green_fill
                current_row += 1
            ws.column_dimensions['A'].width=20; ws.column_dimensions['B'].width=12; ws.column_dimensions['C'].width=10; ws.column_dimensions['D'].width=10; ws.column_dimensions['E'].width=10; ws.column_dimensions['F'].width=10; ws.column_dimensions['G'].width=10; ws.column_dimensions['H'].width=45; ws.column_dimensions['I'].width=10; ws.column_dimensions['J'].width=10; ws.column_dimensions['K'].width=10; ws.column_dimensions['L'].width=10
            wb.save(filepath); messagebox.showinfo("Export Successful", f"Batch results exported to:\n{filepath}"); self.set_status("Batch Excel Export Successful.")
        except PermissionError: messagebox.showerror("Export Error", f"Permission denied: {filepath}"); self.set_status("Error: Batch Excel Permission Denied.")
        except Exception as e: messagebox.showerror("Excel Export Error", f"Error creating batch Excel:\n{e}"); log_event("ERROR", details=f"Batch Excel Export failed: {e}\n{traceback.format_exc()}"); self.set_status("Error: Batch Excel Export Failed.")

    # --- Worklist Methods ---
    def show_worklist_window(self):
        """Creates and shows the worklist manager window."""
        work_win = tk.Toplevel(self.master); work_win.title("Worklist Manager (Session Only)"); work_win.geometry("400x500"); work_win.transient(self.master)
        controls_frame = ttk.Frame(work_win, padding="10"); controls_frame.pack(fill=tk.X, side=tk.TOP)
        ttk.Label(controls_frame, text="Sample ID:").pack(side=tk.LEFT, padx=(0, 5)); add_entry = ttk.Entry(controls_frame, width=20); add_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True); add_button = ttk.Button(controls_frame, text="Add", width=5, command=lambda: self._worklist_add(add_entry, listbox)); add_button.pack(side=tk.LEFT, padx=5); add_entry.bind("<Return>", lambda event: self._worklist_add(add_entry, listbox))
        list_frame = ttk.Frame(work_win, padding="10"); list_frame.pack(fill=tk.BOTH, expand=True)
        listbox = tk.Listbox(list_frame, height=15, selectmode=tk.SINGLE); listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True);
        for item in self.worklist_items: listbox.insert(tk.END, item)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=listbox.yview); scrollbar.pack(side=tk.RIGHT, fill=tk.Y); listbox.config(yscrollcommand=scrollbar.set)
        action_frame = ttk.Frame(work_win, padding="10"); action_frame.pack(fill=tk.X, side=tk.BOTTOM)
        load_button = ttk.Button(action_frame, text="Load Selected Sample", command=lambda: self._worklist_load(listbox, work_win)); load_button.pack(side=tk.LEFT, padx=5)
        remove_button = ttk.Button(action_frame, text="Remove Selected", command=lambda: self._worklist_remove(listbox)); remove_button.pack(side=tk.LEFT, padx=5)
        close_button = ttk.Button(action_frame, text="Close", command=work_win.destroy); close_button.pack(side=tk.RIGHT, padx=5)
        work_win.grab_set(); work_win.focus_set(); work_win.wait_window()

    def _worklist_add(self, entry_widget, listbox_widget):
        """Adds Sample ID from entry to worklist."""
        sample_id = entry_widget.get().strip()
        if sample_id:
            if sample_id not in self.worklist_items: self.worklist_items.append(sample_id); listbox_widget.insert(tk.END, sample_id); entry_widget.delete(0, tk.END); self.set_status(f"'{sample_id}' added to worklist.")
            else: messagebox.showwarning("Duplicate", f"'{sample_id}' already in worklist.", parent=entry_widget.winfo_toplevel())
        entry_widget.focus_set()

    def _worklist_remove(self, listbox_widget):
        """Removes selected item from worklist."""
        selected_indices = listbox_widget.curselection()
        if selected_indices: index = selected_indices[0]; sample_id = listbox_widget.get(index); listbox_widget.delete(index);
        if sample_id in self.worklist_items: self.worklist_items.remove(sample_id); self.set_status(f"'{sample_id}' removed from worklist.")
        else: messagebox.showwarning("Selection Error", "Please select a Sample ID to remove.", parent=listbox_widget.winfo_toplevel())

    def _worklist_load(self, listbox_widget, worklist_window):
        """Loads selected Sample ID into main window."""
        selected_indices = listbox_widget.curselection()
        if selected_indices: index = selected_indices[0]; sample_id = listbox_widget.get(index); self.clear_fields(); self.sample_id_entry.insert(0, sample_id); self.nil_entry.focus_set(); self.set_status(f"Loaded Sample ID '{sample_id}' from worklist."); worklist_window.destroy()
        else: messagebox.showwarning("Selection Error", "Please select a Sample ID to load.", parent=listbox_widget.winfo_toplevel())

    # --- Log Viewer Window ---
    def show_log_viewer_window(self):
        """Creates and shows the log viewer window."""
        log_win = tk.Toplevel(self.master); log_win.title("Interpretation Log Viewer"); log_win.geometry("1100x650"); log_win.transient(self.master)
        controls_frame = ttk.Frame(log_win, padding="10"); controls_frame.pack(fill=tk.X, side=tk.TOP)
        ttk.Label(controls_frame, text="Sample ID:").grid(row=0, column=0, padx=(0,2),pady=5,sticky=tk.W); log_sample_entry = ttk.Entry(controls_frame, width=15); log_sample_entry.grid(row=0, column=1, padx=(0,10),pady=5,sticky=tk.W)
        ttk.Label(controls_frame, text="Run ID:").grid(row=0, column=2, padx=(0,2),pady=5,sticky=tk.W); log_run_entry = ttk.Entry(controls_frame, width=15); log_run_entry.grid(row=0, column=3, padx=(0,10),pady=5,sticky=tk.W)
        ttk.Label(controls_frame, text="Operator ID:").grid(row=0, column=4, padx=(0,2),pady=5,sticky=tk.W); log_op_entry = ttk.Entry(controls_frame, width=10); log_op_entry.grid(row=0, column=5, padx=(0,10),pady=5,sticky=tk.W)
        ttk.Label(controls_frame, text="Date (YYYY-MM-DD):").grid(row=0, column=6, padx=(0,2),pady=5,sticky=tk.W); log_date_entry = ttk.Entry(controls_frame, width=12); log_date_entry.grid(row=0, column=7, padx=(0,10),pady=5,sticky=tk.W)
        log_search_button = ttk.Button(controls_frame, text="Search / Filter Log", command=lambda: load_log_data(tv, log_sample_entry.get(), log_run_entry.get(), log_op_entry.get(), log_date_entry.get())); log_search_button.grid(row=0, column=8, padx=5, pady=5)
        log_refresh_button = ttk.Button(controls_frame, text="Refresh Log", command=lambda: load_log_data(tv, log_sample_entry.get(), log_run_entry.get(), log_op_entry.get(), log_date_entry.get())); log_refresh_button.grid(row=0, column=9, padx=5, pady=5)
        log_close_button = ttk.Button(controls_frame, text="Close", command=log_win.destroy); log_close_button.grid(row=0, column=10, padx=(20,5), pady=5)
        tree_frame = ttk.Frame(log_win, padding="10"); tree_frame.pack(fill=tk.BOTH, expand=True)
        log_columns_ids = [h.lower().replace(' ', '_') for h in LOG_HEADER]; tv = ttk.Treeview(tree_frame, columns=log_columns_ids, show='headings', height=25)
        for i, col_id in enumerate(log_columns_ids): tv.heading(col_id, text=LOG_HEADER[i], anchor=tk.W)
        tv.column("timestamp", width=140, stretch=tk.NO, anchor=tk.W); tv.column("operatorid", width=70, stretch=tk.NO, anchor=tk.W); tv.column("runid", width=100, stretch=tk.NO, anchor=tk.W); tv.column("sampleid", width=120, stretch=tk.NO, anchor=tk.W)
        tv.column("nil", width=60, stretch=tk.NO, anchor=tk.E); tv.column("tb1", width=60, stretch=tk.NO, anchor=tk.E); tv.column("tb2", width=60, stretch=tk.NO, anchor=tk.E); tv.column("mitogen", width=60, stretch=tk.NO, anchor=tk.E)
        tv.column("result", width=70, stretch=tk.NO, anchor=tk.W); tv.column("reason", width=300, stretch=tk.YES, anchor=tk.W)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tv.yview); hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tv.xview); tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tv.grid(row=0, column=0, sticky='nsew'); vsb.grid(row=0, column=1, sticky='ns'); hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1); tree_frame.grid_columnconfigure(0, weight=1)
        load_log_data(tv, "", "", "", ""); log_win.grab_set(); log_win.focus_set(); log_win.wait_window()

    # --- Self-Test Method ---
    def run_self_test(self):
        """Runs predefined test cases against the interpretation logic."""
        self.set_status("Running Self-Test...")
        results_text = "--- QFT Interpreter Self-Test Results ---\n\n"; passed_count = 0; failed_count = 0
        for i, (desc, nil, tb1, tb2, mit, expected) in enumerate(SELF_TEST_CASES):
            test_num = i + 1
            try:
                result_dict = interpret_qft(nil, tb1, tb2, mit); actual = result_dict['result']
                if actual == expected: status = "PASSED"; passed_count += 1
                else: status = f"FAILED (Expected: {expected}, Got: {actual})"; failed_count += 1
                results_text += f"Test {test_num:<2}: {desc:<40} {status}\n"
            except Exception as e: status = f"ERROR ({e})"; failed_count += 1; results_text += f"Test {test_num:<2}: {desc:<40} {status}\n"
        results_text += f"\n--- Summary ---\nPassed: {passed_count}\nFailed/Error: {failed_count}\nTotal: {len(SELF_TEST_CASES)}\n"; log_event("INFO", details=f"Self-Test Executed: {passed_count} Passed, {failed_count} Failed/Error."); self.set_status("Self-Test Complete.")
        test_win = tk.Toplevel(self.master); test_win.title("Self-Test Results"); test_win.geometry("600x400"); test_win.transient(self.master)
        text_area = tk.Text(test_win, wrap=tk.WORD, font=('Consolas', 9), relief=tk.SOLID, borderwidth=1); text_area.pack(padx=10, pady=10, fill=tk.BOTH, expand=True); text_area.insert(tk.END, results_text); text_area.config(state='disabled')
        close_button = ttk.Button(test_win, text="Close", command=test_win.destroy); close_button.pack(pady=5); test_win.grab_set(); test_win.focus_set(); test_win.wait_window()

    # --- Dashboard Methods ---
    def update_dashboard(self):
        """Queries the DB and updates the dashboard labels."""
        self.set_status("Updating dashboard...")
        try:
            days_range = int(self.config.get('dashboard_days', 7)); end_date = datetime.now(); start_date = end_date - timedelta(days=days_range - 1)
            start_date_str = start_date.strftime('%Y-%m-%d'); end_date_str = end_date.strftime('%Y-%m-%d')
            data = query_db_for_reports(start_date_str, end_date_str)
            total = len(data); pos_count = 0; neg_count = 0; ind_count = 0; ind_reasons = Counter()
            for row in data:
                result = row[4]; reason = row[5]
                if result == "POS†": pos_count += 1
                elif result == "NEG": neg_count += 1
                elif result == "IND*":
                    ind_count += 1
                    if "High Nil" in reason: ind_reasons["High Nil"] += 1
                    elif "Low Mitogen" in reason: ind_reasons["Low Mitogen"] += 1
                    else: ind_reasons["Other"] += 1
            pos_rate = (pos_count / total * 100) if total > 0 else 0; ind_rate = (ind_count / total * 100) if total > 0 else 0
            self.dashboard_vars["total"].set(f"Total: {total}"); self.dashboard_vars["pos"].set(f"POS: {pos_count}"); self.dashboard_vars["neg"].set(f"NEG: {neg_count}"); self.dashboard_vars["ind"].set(f"IND: {ind_count}")
            self.dashboard_vars["ind_high_nil"].set(f"IND (High Nil): {ind_reasons['High Nil']}"); self.dashboard_vars["ind_low_mit"].set(f"IND (Low Mitogen): {ind_reasons['Low Mitogen']}")
            self.dashboard_vars["pos_rate"].set(f"POS%: {pos_rate:.1f}%"); self.dashboard_vars["ind_rate"].set(f"IND%: {ind_rate:.1f}%")
            self.dashboard_frame.config(text=f"Dashboard ({start_date_str} to {end_date_str} - {days_range} Days)")
            self.set_status("Dashboard updated."); # No separate log event for dashboard update
        except Exception as e:
            self.set_status("Error updating dashboard."); log_event("ERROR", details=f"Dashboard update failed: {e}\n{traceback.format_exc()}")
            for key in self.dashboard_vars: self.dashboard_vars[key].set(f"{key.replace('_',' ').title()}: Error")

    # --- Report Generation Window ---
    def show_report_window(self):
        """Shows the window for generating summary reports."""
        report_win = tk.Toplevel(self.master); report_win.title("Generate Summary Report"); report_win.geometry("400x200"); report_win.transient(self.master)
        frame = ttk.Frame(report_win, padding="15"); frame.pack(expand=True, fill=tk.BOTH)
        ttk.Label(frame, text="Select Date Range:").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0,5))
        ttk.Label(frame, text="Start Date (YYYY-MM-DD):").grid(row=1, column=0, sticky="w", padx=5, pady=2); start_date_entry = ttk.Entry(frame, width=12); start_date_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=2)
        today = datetime.now(); default_start = today - timedelta(days=self.config.get('dashboard_days', 7) - 1); start_date_entry.insert(0, default_start.strftime('%Y-%m-%d'))
        ttk.Label(frame, text="End Date (YYYY-MM-DD):").grid(row=2, column=0, sticky="w", padx=5, pady=2); end_date_entry = ttk.Entry(frame, width=12); end_date_entry.grid(row=2, column=1, sticky="ew", padx=5, pady=2); end_date_entry.insert(0, today.strftime('%Y-%m-%d'))
        button_frame = ttk.Frame(frame, padding="0 15 0 0"); button_frame.grid(row=3, column=0, columnspan=2, sticky="e")
        gen_pdf_button = ttk.Button(button_frame, text="Generate PDF Report", command=lambda: self._generate_report('pdf', start_date_entry.get(), end_date_entry.get()), state=(tk.NORMAL if REPORTLAB_AVAILABLE else tk.DISABLED)); gen_pdf_button.pack(side=tk.RIGHT, padx=5)
        gen_excel_button = ttk.Button(button_frame, text="Generate Excel Report", command=lambda: self._generate_report('excel', start_date_entry.get(), end_date_entry.get()), state=(tk.NORMAL if OPENPYXL_AVAILABLE else tk.DISABLED)); gen_excel_button.pack(side=tk.RIGHT, padx=5)
        report_win.grab_set(); report_win.focus_set(); report_win.wait_window()

    def _generate_report(self, format_type, start_date_str, end_date_str):
        """Fetches data and generates the summary report."""
        self.set_status(f"Generating {format_type.upper()} report..."); data = query_db_for_reports(start_date_str, end_date_str)
        if data is None: self.set_status("Report generation failed (DB query error)."); return
        if not data: messagebox.showinfo("No Data", f"No records found between {start_date_str} and {end_date_str}.", parent=self.master); self.set_status("Report generation cancelled (no data)."); return
        total = len(data); pos_count = 0; neg_count = 0; ind_count = 0; ind_reasons = Counter(); run_ids = set(); operators = set();
        for row in data:
            timestamp, operator, run_id, sample_id, result, reason = row
            if operator: operators.add(operator)
            if run_id: run_ids.add(run_id)
            if result == "POS†": pos_count += 1
            elif result == "NEG": neg_count += 1
            elif result == "IND*": ind_count += 1;
            if "High Nil" in reason: ind_reasons["High Nil"] += 1
            elif "Low Mitogen" in reason: ind_reasons["Low Mitogen"] += 1
            else: ind_reasons["Other"] += 1
        pos_rate=(pos_count/total*100) if total>0 else 0; neg_rate=(neg_count/total*100) if total>0 else 0; ind_rate=(ind_count/total*100) if total>0 else 0;
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S"); default_filename = f"QFT_Summary_Report_{start_date_str}_to_{end_date_str}_{timestamp}"
        if format_type == 'pdf': file_ext=".pdf"; file_types=[("PDF Documents", "*.pdf")]
        elif format_type == 'excel': file_ext=".xlsx"; file_types=[("Excel Workbook", "*.xlsx")]
        else: return
        filepath = filedialog.asksaveasfilename(defaultextension=file_ext, filetypes=file_types, initialfile=default_filename, title=f"Save Summary Report as {format_type.upper()}")
        if not filepath: self.set_status("Report export cancelled."); return
        success = False
        try:
            if format_type == 'pdf': success = self._write_summary_pdf(filepath, start_date_str, end_date_str, total, pos_count, neg_count, ind_count, pos_rate, neg_rate, ind_rate, ind_reasons, run_ids, operators)
            elif format_type == 'excel': success = self._write_summary_excel(filepath, start_date_str, end_date_str, total, pos_count, neg_count, ind_count, pos_rate, neg_rate, ind_rate, ind_reasons, run_ids, operators, data)
            if success: messagebox.showinfo("Report Generated", f"Summary report saved successfully to:\n{filepath}"); self.set_status(f"{format_type.upper()} Report Generated.")
        except Exception as e: messagebox.showerror("Report Generation Error", f"Failed to generate {format_type.upper()} report:\n{e}"); log_event("ERROR", details=f"Report Generation ({format_type}) failed: {e}\n{traceback.format_exc()}"); self.set_status(f"Error generating {format_type.upper()} report.")

    def _write_summary_pdf(self, filepath, start_date, end_date, total, pos, neg, ind, pos_r, neg_r, ind_r, ind_reasons, runs, ops):
        """Helper to write the summary data to a PDF file."""
        if not REPORTLAB_AVAILABLE: return False
        try:
            doc = SimpleDocTemplate(filepath, pagesize=(8.5*inch, 11*inch), leftMargin=0.75*inch, rightMargin=0.75*inch, topMargin=0.75*inch, bottomMargin=0.75*inch); styles = getSampleStyleSheet(); story = []; styles.add(ParagraphStyle(name='RightAlign', alignment=TA_RIGHT)); styles.add(ParagraphStyle(name='Bold', fontName='Helvetica-Bold'))
            story.append(Paragraph("QFT Interpretation Summary Report", styles['h1'])); story.append(Paragraph(f"Date Range: {start_date} to {end_date}", styles['h3'])); story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])); story.append(Spacer(1, 0.3*inch))
            summary_data = [[Paragraph('Metric', styles['Bold']), Paragraph('Count / Value', styles['Bold'])],['Total Interpretations:', Paragraph(str(total), styles['RightAlign'])],['Positive Results (POS†):', Paragraph(str(pos), styles['RightAlign'])],['Negative Results (NEG):', Paragraph(str(neg), styles['RightAlign'])],['Indeterminate Results (IND*):', Paragraph(str(ind), styles['RightAlign'])],['    - IND (High Nil):', Paragraph(str(ind_reasons.get("High Nil", 0)), styles['RightAlign'])],['    - IND (Low Mitogen):', Paragraph(str(ind_reasons.get("Low Mitogen", 0)), styles['RightAlign'])],['    - IND (Other):', Paragraph(str(ind_reasons.get("Other", 0)), styles['RightAlign'])],['Positivity Rate:', Paragraph(f"{pos_r:.1f}%", styles['RightAlign'])],['Negativity Rate:', Paragraph(f"{neg_r:.1f}%", styles['RightAlign'])],['Indeterminate Rate:', Paragraph(f"{ind_r:.1f}%", styles['RightAlign'])],['Unique Run IDs:', Paragraph(str(len(runs)), styles['RightAlign'])],['Unique Operators:', Paragraph(str(len(ops)), styles['RightAlign'])]]
            table = Table(summary_data, colWidths=[3*inch, 1.5*inch]); table.setStyle(TableStyle([('GRID',(0,0),(-1,-1),1,colors.black), ('BACKGROUND',(0,0),(-1,0),colors.grey), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke), ('ALIGN',(0,0),(-1,0),'CENTER'), ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('BOTTOMPADDING',(0,0),(-1,-1),6), ('TOPPADDING',(0,0),(-1,-1),6), ('ALIGN',(0,1),(0,-1),'LEFT')])); story.append(table)
            doc.build(story); return True
        except Exception as e: messagebox.showerror("PDF Write Error", f"Error creating PDF:\n{e}"); log_event("ERROR", details=f"Summary PDF Write failed: {e}\n{traceback.format_exc()}"); return False

    def _write_summary_excel(self, filepath, start_date, end_date, total, pos, neg, ind, pos_r, neg_r, ind_r, ind_reasons, runs, ops, detail_data):
        """Helper to write the summary data and detail to an Excel file."""
        if not OPENPYXL_AVAILABLE: return False
        try:
            wb = Workbook(); ws_summary = wb.active; ws_summary.title = "Summary"; bold_font=Font(bold=True); right_align=Alignment(horizontal='right')
            ws_summary['A1'] = "QFT Interpretation Summary Report"; ws_summary['A1'].font = Font(bold=True, size=14); ws_summary['A2'] = "Date Range:"; ws_summary['B2'] = f"{start_date} to {end_date}"; ws_summary['A3'] = "Report Generated:"; ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            summary_headers = ["Metric", "Count / Value"]; summary_rows = [("Total Interpretations:", total), ("Positive Results (POS†):", pos), ("Negative Results (NEG):", neg), ("Indeterminate Results (IND*):", ind), ("    - IND (High Nil):", ind_reasons.get("High Nil", 0)), ("    - IND (Low Mitogen):", ind_reasons.get("Low Mitogen", 0)), ("    - IND (Other):", ind_reasons.get("Other", 0)), ("Positivity Rate:", f"{pos_r:.1f}%"), ("Negativity Rate:", f"{neg_r:.1f}%"), ("Indeterminate Rate:", f"{ind_r:.1f}%"), ("Unique Run IDs:", len(runs)), ("Unique Operators:", len(ops))]
            current_row = 5
            for col_idx, header in enumerate(summary_headers, 1): ws_summary.cell(row=current_row, column=col_idx, value=header).font = bold_font
            current_row += 1
            for metric, value in summary_rows: ws_summary.cell(row=current_row, column=1, value=metric); cell = ws_summary.cell(row=current_row, column=2, value=value); cell.alignment = right_align;
            if isinstance(value, (int, float)): cell.number_format = '0'
            else: cell.number_format = '@'; current_row += 1
            ws_summary.column_dimensions['A'].width = 30; ws_summary.column_dimensions['B'].width = 15
            ws_detail = wb.create_sheet("Detail"); detail_headers = ["Timestamp", "Operator", "Run ID", "Sample ID", "Result", "Reason"]
            for col_idx, header in enumerate(detail_headers, 1): ws_detail.cell(row=1, column=col_idx, value=header).font = bold_font
            for row_idx, data_row in enumerate(detail_data, 2):
                 for col_idx, cell_value in enumerate(data_row, 1): ws_detail.cell(row=row_idx, column=col_idx, value=cell_value)
            ws_detail.column_dimensions['A'].width = 20; ws_detail.column_dimensions['B'].width = 15; ws_detail.column_dimensions['C'].width = 15; ws_detail.column_dimensions['D'].width = 20; ws_detail.column_dimensions['E'].width = 10; ws_detail.column_dimensions['F'].width = 50
            wb.save(filepath); return True
        except Exception as e: messagebox.showerror("Excel Write Error", f"Error creating Excel:\n{e}"); log_event("ERROR", details=f"Summary Excel Write failed: {e}\n{traceback.format_exc()}"); return False

    # --- Window Closing ---
    def on_closing(self):
        """Handles window closing: saves config."""
        current_geometry = self.master.geometry(); self.config['geometry'] = current_geometry; self.config['theme'] = self.current_theme.get()
        save_config(self.config); log_event("INFO", details="Application Shutdown.")
        self.master.destroy()


# --- History & Log Loading Functions (outside class) ---
def load_history(treeview, search_id=None, search_date=None, search_run_id=None):
    """Loads data into the history Treeview, optionally filtering."""
    for item in treeview.get_children(): treeview.delete(item); conn = None
    try:
        conn = sqlite3.connect(DB_FILENAME); cursor = conn.cursor(); query = "SELECT timestamp, operator_id, run_id, sample_id, nil_value, tb1_value, tb2_value, mit_value, result, reason FROM interpretations"; params = []; conditions = []
        effective_search_id = search_id.strip() if search_id else None; effective_search_date = search_date.strip() if search_date else None; effective_search_run_id = search_run_id.strip() if search_run_id else None
        if effective_search_id: conditions.append("sample_id LIKE ?"); params.append(f"%{effective_search_id}%")
        if effective_search_run_id: conditions.append("run_id LIKE ?"); params.append(f"%{effective_search_run_id}%")
        if effective_search_date:
            try: datetime.strptime(effective_search_date, '%Y-%m-%d'); conditions.append("DATE(timestamp) = ?"); params.append(effective_search_date)
            except ValueError: messagebox.showerror("Invalid Date", "Use YYYY-MM-DD format.", parent=treeview.winfo_toplevel()); conn.close(); return
        if conditions: query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY timestamp DESC LIMIT 500"; cursor.execute(query, params); rows = cursor.fetchall()
        if not rows:
             if effective_search_id or effective_search_date or effective_search_run_id: msg = "No records found matching filters."
             else: msg = "No history records found."
             treeview.insert('', tk.END, values=(msg, *[""]*9 ))
        else:
            for row in rows:
                 formatted_row = list(row);
                 for i in [4, 5, 6, 7]:
                    try: formatted_row[i] = f"{float(formatted_row[i]):.3f}" if formatted_row[i] is not None else ""
                    except: formatted_row[i] = str(formatted_row[i])
                 treeview.insert('', tk.END, values=tuple(formatted_row))
    except sqlite3.Error as e: messagebox.showerror("Database Error", f"Failed to load history:\n{e}", parent=treeview.winfo_toplevel())
    except Exception as e: traceback.print_exc(); messagebox.showerror("History Error", f"Error loading history:\n{e}", parent=treeview.winfo_toplevel())
    finally:
        if conn: conn.close()

# --- History & Log Loading Functions (outside class) ---
# (load_history function would be here too)

def load_log_data(treeview, search_sample="", search_run="", search_op="", search_date=""):
    """Loads and filters data from the CSV log file into the log viewer Treeview."""
    for item in treeview.get_children(): treeview.delete(item)
    rows_loaded = 0 # Initialize rows_loaded count here
    try:
        # Check if file exists before trying to open
        if not os.path.exists(LOG_FILENAME):
             messagebox.showinfo("Log File", "Log file not found. It will be created when the first interpretation is logged.", parent=treeview.winfo_toplevel())
             treeview.insert('', tk.END, values=("Log file not found.", *[""]*(len(LOG_HEADER)-1) ))
             return # Exit function early if file doesn't exist

        # File exists, proceed to read
        with open(LOG_FILENAME, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            header = next(reader, None) # Read/skip header

            # Check header length and content (optional warning)
            if not header or len(header) != len(LOG_HEADER) or [h.lower() for h in header] != [h.lower() for h in LOG_HEADER]:
                 messagebox.showwarning("Log Format Warning", f"Log header/column count mismatch ({len(header) if header else 'None'} found, {len(LOG_HEADER)} expected). Display might be incorrect.", parent=treeview.winfo_toplevel())
                 # Attempt to display anyway? Might be risky if columns shifted.

            filtered_rows = []
            row_num_debug = 1 # Start counting after header for debug messages
            for row in reader:
                row_num_debug += 1
                # Check column count rigorously for data rows
                if len(row) != len(LOG_HEADER):
                    print(f"Warning: Skipping log row {row_num_debug} due to incorrect column count ({len(row)} found, {len(LOG_HEADER)} expected).", file=sys.stderr)
                    continue # Skip malformed rows

                # Apply filters (case-insensitive for strings)
                # Unpack based on the 10-column LOG_HEADER order
                timestamp_str, op_id, run_id, sample_id, nil_val_str, tb1_val_str, tb2_val_str, mit_val_str, result_val, reason_val = row

                match = True
                if search_sample and search_sample.lower() not in sample_id.lower(): match = False
                if search_run and search_run.lower() not in run_id.lower(): match = False
                if search_op and search_op.lower() not in op_id.lower(): match = False
                if search_date:
                    try:
                        # Attempt to parse date part of timestamp
                        log_date_str = datetime.strptime(timestamp_str.split()[0], '%Y-%m-%d').strftime('%Y-%m-%d')
                        if log_date_str != search_date: match = False
                    except (ValueError, IndexError):
                        # Handle cases where timestamp might be malformed or date parsing fails
                        print(f"Warning: Could not parse date from log timestamp '{timestamp_str}' for filtering.", file=sys.stderr)
                        match = False # Treat as non-match if date filter applied and parsing fails

                if match:
                    filtered_rows.append(row) # Keep matching rows

            # Insert filtered rows (most recent first)
            for row_data in reversed(filtered_rows): # Show recent logs first
                 treeview.insert('', tk.END, values=tuple(row_data))
                 rows_loaded += 1

            if rows_loaded == 0:
                 # Check if filters were applied to distinguish between "no match" and "empty file"
                 if search_sample or search_run or search_op or search_date:
                     message = "No log entries found matching current filters."
                 else:
                     message = "Log file is empty or contains no valid entries."
                 treeview.insert('', tk.END, values=(message, *[""]*(len(LOG_HEADER)-1) ))

    # Keep FileNotFoundError separate as it's handled differently now
    # except FileNotFoundError: handled above

    except Exception as e:
        # Generic error during file reading/processing
        # Use the log_event helper if available, otherwise print
        try:
            log_event("ERROR", details=f"Log Viewer read/parse error: {e}\n{traceback.format_exc()}")
        except NameError: # If log_event isn't defined globally (should be, but safe)
             print(f"ERROR: Log Viewer read/parse error: {e}\n{traceback.format_exc()}", file=sys.stderr)

        messagebox.showerror("Log Viewer Error", f"Failed to read or parse log file:\n{e}", parent=treeview.winfo_toplevel())
        # Optionally display error in treeview too
        treeview.insert('', tk.END, values=(f"Error loading log: {e}", *[""]*(len(LOG_HEADER)-1) ))

# --- Main Execution ---
if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = QFTApp(root)
        root.mainloop()
    except Exception as e:
        try: log_event("CRITICAL", details=f"GUI Startup Failed: {e}\n{traceback.format_exc()}")
        except: pass
        try: messagebox.showerror("Fatal Error", f"Application failed to start:\n{e}\n\nPlease check dependencies and logs.")
        except: pass
        print(f"CRITICAL ERROR: Application failed to start: {e}", file=sys.stderr)
        traceback.print_exc()
        sys.exit(1)